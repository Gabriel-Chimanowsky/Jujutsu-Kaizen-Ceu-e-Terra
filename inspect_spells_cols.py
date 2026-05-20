import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet = wb['Perfil Amaldiçoado']

print("--- Row 8 merged status / cells ---")
for c in range(27, 54): # AA is 27, AS is 45, BB is 54
    val = sheet.cell(row=8, column=c).value
    # Let's print the cell value and coordinate
    coord = sheet.cell(row=8, column=c).coordinate
    print(f"Col {c} ({coord}): {val}")
