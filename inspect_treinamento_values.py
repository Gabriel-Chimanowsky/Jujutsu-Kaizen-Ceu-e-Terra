import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet = wb['Treinamentos']

print("--- TREINAMENTO DE AGILIDADE CELLS ---")
for r in range(5, 10):
    for c in range(3, 11): # C is 3, J is 10
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            coord = sheet.cell(row=r, column=c).coordinate
            print(f"[{coord}] = {val}")
