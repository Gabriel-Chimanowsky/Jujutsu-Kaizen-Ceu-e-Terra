import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet = wb['Treinamentos']

print("--- ALL TREINAMENTOS ---")
for r in [4, 16]:
    for c in range(3, 50, 9): # C(3), L(12), U(21), AD(30), AM(39), AV(48)
        title = sheet.cell(row=r, column=c).value
        if title:
            # Let's read the description at row r+6 (which is row 10 or 22)
            desc_row = r + 6
            desc = sheet.cell(row=desc_row, column=c).value
            print(f"- {title} (Row {r}, Col {c}): {str(desc)[:100]}...")
