import openpyxl
wb = openpyxl.load_workbook(r'exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx', data_only=True)

# Perfil with smaller stride
ws2 = wb['Perfil Amaldiçoado']
print('=== PERFIL AMALDICADO - spell areas (rows 5-30) ===')
for row in range(5, 35):
    for col in range(26, 80):
        v = ws2.cell(row=row, column=col).value
        if v is not None:
            print('  ' + ws2.cell(row=row, column=col).column_letter + str(row) + ': ' + str(v))

print()
ws_t = wb['Treinamentos']
print('=== TREINAMENTOS all content ===')
for row in range(1, 38):
    for col in range(1, 58, 2):
        v = ws_t.cell(row=row, column=col).value
        if v is not None:
            print('  ' + ws_t.cell(row=row, column=col).column_letter + str(row) + ': ' + str(v))

# Inventory with item columns
print()
ws_i = wb['Registro e Inventário']
print('=== INVENTÁRIO detailed item rows (rows 6-35) ===')
for row in range(6, 36):
    for col in range(24, 54):
        v = ws_i.cell(row=row, column=col).value
        if v is not None:
            print('  ' + ws_i.cell(row=row, column=col).column_letter + str(row) + ': ' + str(v))
