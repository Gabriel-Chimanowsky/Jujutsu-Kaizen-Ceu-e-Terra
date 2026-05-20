import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet = wb['Ficha Pessoal']

print("--- FICHA TOP ROWS (Cols O-U, Rows 1-16) ---")
for r in range(1, 17):
    row_vals = []
    for c in range(14, 25): # N(14) to X(24)
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            coord = sheet.cell(row=r, column=c).coordinate
            row_vals.append(f"{coord}:{val}")
    if row_vals:
        print(f"Row {r}: {row_vals}")
