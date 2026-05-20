import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet = wb['Ficha Pessoal']

print("--- RDs REGION (Cols Z-AE, Rows 20-25) ---")
for r in range(20, 26):
    row_vals = []
    for c in range(26, 32): # Z(26) to AE(31)
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            coord = sheet.cell(row=r, column=c).coordinate
            row_vals.append(f"{coord}:{val}")
    if row_vals:
        print(f"Row {r}: {row_vals}")
