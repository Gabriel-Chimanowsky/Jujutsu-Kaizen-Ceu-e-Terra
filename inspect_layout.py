import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

# 1. Ficha Pessoal - Attacks Area
print("=== FICHA PESSOAL: ATAQUES (AA25 to AU36) ===")
sheet1 = wb['Ficha Pessoal']
for r in range(25, 36):
    row_vals = []
    for c in range(27, 45): # col AA to AR
        val = sheet1.cell(row=r, column=c).value
        cell_coord = sheet1.cell(row=r, column=c).coordinate
        if val is not None:
            row_vals.append(f"{cell_coord}:{val}")
    if row_vals:
        print(f"Row {r}: {row_vals}")

# 2. Registro e Inventário - Inventory Area
print("\n=== REGISTRO E INVENTARIO: INVENTARIO (col Y to AX, row 5 to 36) ===")
sheet3 = wb['Registro e Inventário']
for r in range(5, 36):
    row_vals = []
    for c in range(25, 54): # col Y to BB
        val = sheet3.cell(row=r, column=c).value
        cell_coord = sheet3.cell(row=r, column=c).coordinate
        if val is not None:
            row_vals.append(f"{cell_coord}:{val}")
    if row_vals:
        print(f"Row {r}: {row_vals}")

# 3. Perfil Amaldiçoado - Spells Area
print("\n=== PERFIL AMALDIÇOADO: FEITIÇOS (col AA to BH, row 5 to 36) ===")
sheet4 = wb['Perfil Amaldiçoado']
for r in range(5, 36):
    row_vals = []
    for c in range(27, 60): # col AA to BH
        val = sheet4.cell(row=r, column=c).value
        cell_coord = sheet4.cell(row=r, column=c).coordinate
        if val is not None:
            row_vals.append(f"{cell_coord}:{val}")
    if row_vals:
        print(f"Row {r}: {row_vals}")
