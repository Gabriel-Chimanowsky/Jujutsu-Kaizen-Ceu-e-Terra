import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx")
sheet = wb["Ficha Pessoal"]

print("=== Merged Ranges overlapping Rows 8 or 9 ===")
for r in sheet.merged_cells.ranges:
    if r.min_row <= 9 <= r.max_row or r.min_row <= 8 <= r.max_row:
        print(f"Range: {r} | Top-Left cell: {r.start_cell}")
