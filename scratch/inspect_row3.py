import openpyxl

wb_v = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet_v = wb_v['Ficha Pessoal']

wb_f = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet_f = wb_f['Ficha Pessoal']

print("--- ROW 3 ---")
for c in range(1, 100):
    val_v = sheet_v.cell(row=3, column=c).value
    val_f = sheet_f.cell(row=3, column=c).value
    if val_v is not None or val_f is not None:
        print(f"Col {openpyxl.utils.get_column_letter(c)}3: Val={val_v} | Form={val_f}")
