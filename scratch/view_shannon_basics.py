import openpyxl

wb = openpyxl.load_workbook("exemplo/Ficha F&M Shannon 2.5.2.xlsx", data_only=True)

print("=== Scanning entire workbook for 'shannon' ===")
found = False
for name in wb.sheetnames:
    ws = wb[name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                v_str = str(v).lower()
                if 'shannon' in v_str:
                    print(f"[{name}] {openpyxl.utils.get_column_letter(c)}{r}: {v}")
                    found = True
if not found:
    print("No cell containing 'shannon' was found.")
