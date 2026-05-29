import openpyxl

wb_v = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet_v = wb_v['Ficha Pessoal']

wb_f = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet_f = wb_f['Ficha Pessoal']

print("--- INSPECTING PE CELL BLOCK ---")
for r in range(6, 13):
    row_str = []
    for c in range(29, 39): # AC (29) to AH (38)
        col_let = openpyxl.utils.get_column_letter(c)
        v = sheet_v.cell(row=r, column=c).value
        f = sheet_f.cell(row=r, column=c).value
        # Check if cell is in merged ranges
        is_merged = False
        for rng in sheet_v.merged_cells.ranges:
            if r >= rng.min_row and r <= rng.max_row and c >= rng.min_col and c <= rng.max_col:
                is_merged = True
                top_left = openpyxl.utils.get_column_letter(rng.min_col) + str(rng.min_row)
                break
        merged_info = f" (Merged, TL={top_left})" if is_merged else ""
        if v is not None or f is not None or is_merged:
            row_str.append(f"{col_let}{r}: v={v}/f={f}{merged_info}")
    if row_str:
        print(f"Row {r}: {', '.join(row_str)}")
