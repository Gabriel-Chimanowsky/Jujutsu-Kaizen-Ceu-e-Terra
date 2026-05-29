import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet = wb["Ficha Pessoal"]

for r in range(1, 100):
    for c in range(1, 100):
        cell = sheet.cell(row=r, column=c)
        val = str(cell.value or "")
        if "PE" in val or "Estamina" in val or "Energia" in val:
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"{col_letter}{r}: Value/Formula={val}")
