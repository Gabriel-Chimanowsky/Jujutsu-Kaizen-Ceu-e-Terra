import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
ws_f = wb_formula['Ficha Pessoal']

for col in range(14, 25):
    val_f = ws_f.cell(row=30, column=col).value
    col_let = openpyxl.utils.get_column_letter(col)
    if val_f is not None:
        print(f"{col_let}30: {val_f}")
