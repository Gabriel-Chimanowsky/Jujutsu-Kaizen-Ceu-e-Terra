import openpyxl

wb = openpyxl.load_workbook("exemplo/Ficha F&M Shannon 2.5.2.xlsx", data_only=True)

print("=== Scanning sheets for any character names ===")
for name in wb.sheetnames:
    ws = wb[name]
    for r in range(1, 40):
        for c in range(1, 40):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                val_str = str(val).strip().lower()
                if any(x in val_str for x in ['shannon', 'yutsuki', 'otokanutti', 'nome', 'character', 'personagem']):
                    print(f"[{name}] {openpyxl.utils.get_column_letter(c)}{r}: {val} -> Next cells: {ws.cell(row=r, column=c+1).value} | {ws.cell(row=r, column=c+2).value}")
