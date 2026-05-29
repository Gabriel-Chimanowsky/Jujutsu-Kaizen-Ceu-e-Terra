import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
ws = wb_formula['Ficha Pessoal']

print("=== Merged Ranges in Ficha Pessoal ===")
for r in sorted(list(ws.merged_cells.ranges), key=lambda x: (x.min_row, x.min_col)):
    # print the range and value of the top-left cell
    top_left_cell = ws.cell(row=r.min_row, column=r.min_col)
    val = top_left_cell.value
    if val is not None:
        col_let = openpyxl.utils.get_column_letter(r.min_col)
        print(f"Range {r.coord}: Value='{val}'")
