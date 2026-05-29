import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n--- Checking sheet {sname} ---")
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).startswith('='):
                val_str = str(val).upper()
                if 'PE' in val_str or 'ENERGIA' in val_str or 'CALCULOS!' in val_str:
                    col_let = openpyxl.utils.get_column_letter(c)
                    print(f"Cell {col_let}{r}: {val}")
