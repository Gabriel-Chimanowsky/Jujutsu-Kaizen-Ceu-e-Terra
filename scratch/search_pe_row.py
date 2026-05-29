import openpyxl

wb_v = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet_v = wb_v['Ficha Pessoal']

wb_f = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet_f = wb_f['Ficha Pessoal']

print("--- SEARCHING COLUMNS AC TO AL (COLS 29 TO 38) IN ROWS 1 TO 25 ---")
for r in range(1, 26):
    for c in range(29, 39):
        val_v = sheet_v.cell(row=r, column=c).value
        val_f = sheet_f.cell(row=r, column=c).value
        if val_v is not None or val_f is not None:
            col_let = openpyxl.utils.get_column_letter(c)
            print(f"Cell {col_let}{r}: Val={val_v} | Form={val_f}")
