import openpyxl

wb_v = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet_v = wb_v['Ficha Pessoal']

wb_f = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet_f = wb_f['Ficha Pessoal']

print("--- ROWS 2 TO 5 ---")
for r in range(2, 6):
    print(f"\n--- Row {r} ---")
    for c in range(1, 100):
        val_v = sheet_v.cell(row=r, column=c).value
        val_f = sheet_f.cell(row=r, column=c).value
        if val_v is not None or val_f is not None:
            print(f"  Col {openpyxl.utils.get_column_letter(c)}{r}: Val={val_v} | Form={val_f}")
