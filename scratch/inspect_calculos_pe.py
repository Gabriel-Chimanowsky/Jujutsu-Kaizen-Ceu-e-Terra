import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
ws = wb['Calculos']

print("--- CALCULOS SHEET CONTENT ---")
for r in range(1, ws.max_row + 1):
    row_str = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_str.append(f"{col_letter}{r}: {v}")
    if row_str:
        print(f"Row {r}: {', '.join(row_str)}")
