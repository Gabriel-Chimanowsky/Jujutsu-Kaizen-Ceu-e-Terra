import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n--- Checking sheet {sname} ---")
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                val_str = str(val).upper()
                if 'ENERGIA' in val_str or ' PE' in val_str or val_str == 'PE' or 'AMALDIÇOADA' in val_str or 'AMALDICOADA' in val_str:
                    col_let = openpyxl.utils.get_column_letter(c)
                    print(f"Cell {col_let}{r}: '{val}'")
