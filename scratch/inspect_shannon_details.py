import openpyxl

wb = openpyxl.load_workbook("exemplo/Ficha F&M Shannon 2.5.2.xlsx", data_only=True)

for name in ['Perfil Mundano', 'Perfil Amaldiçoado']:
    if name in wb.sheetnames:
        ws = wb[name]
        print(f"=== {name} (Rows 1-15, Cols A-M) ===")
        for r in range(1, 16):
            row_vals = []
            for c in range(1, 14):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {v}")
            if row_vals:
                print(" | ".join(row_vals))
