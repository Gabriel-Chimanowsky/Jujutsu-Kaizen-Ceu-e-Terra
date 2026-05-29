import openpyxl

wb = openpyxl.load_workbook('exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx', data_only=True)
sheet = wb['Ficha Pessoal']

# Let's inspect rows 13 to 19 and columns 15 to 24
for r in range(13, 20):
    row_vals = []
    for c in range(15, 25):
        cell_name = openpyxl.utils.get_column_letter(c) + str(r)
        val = sheet.cell(row=r, column=c).value
        row_vals.append(f"{cell_name}: {val}")
    print(f"Row {r} -> " + " | ".join(row_vals))
