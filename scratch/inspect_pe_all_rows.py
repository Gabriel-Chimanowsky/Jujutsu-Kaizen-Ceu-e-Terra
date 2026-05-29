import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

ws_f = wb_formula['Ficha Pessoal']
ws_v = wb_value['Ficha Pessoal']

print("=== Direct Scan for columns AC to AK in rows 1 to 20 ===")
for r in range(1, 21):
    for c in range(29, 38):
        vf = ws_f.cell(row=r, column=c).value
        vv = ws_v.cell(row=r, column=c).value
        if vf is not None or vv is not None:
            col_let = openpyxl.utils.get_column_letter(c)
            print(f"Cell {col_let}{r}: Formula='{vf}' | Value='{vv}'")
