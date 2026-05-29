import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

ws_f = wb_formula['Ficha Pessoal']
ws_v = wb_value['Ficha Pessoal']

print("=== Searching for Key Cells in 'Ficha Pessoal' ===")
for row in range(1, 120):
    for col in range(1, 100):
        val_v = ws_v.cell(row=row, column=col).value
        val_f = ws_f.cell(row=row, column=col).value
        if val_v is not None:
            s = str(val_v).lower()
            if any(term in s for term in ["vida", "pe", "energia", "defesa", "morte", "iniciativa", "espectador", "lutador", "controlador", "suporte", "nível", "nivel"]):
                col_let = openpyxl.utils.get_column_letter(col)
                print(f"Cell {col_let}{row}: Value='{val_v}' | Formula='{val_f}'")
