import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
ws_f = wb_formula['Ficha Pessoal']
ws_v = wb_value['Ficha Pessoal']

for r in range(17, 21):
    for c in range(8, 14):
        val_f = ws_f.cell(row=r, column=c).value
        val_v = ws_v.cell(row=r, column=c).value
        if val_f is not None or val_v is not None:
            col_let = openpyxl.utils.get_column_letter(c)
            print(f"{col_let}{r}: Formula='{val_f}' | Value='{val_v}'")
