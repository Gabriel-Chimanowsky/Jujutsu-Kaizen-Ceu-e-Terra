import openpyxl

wb_v = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet_v = wb_v['Ficha Pessoal']

wb_f = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet_f = wb_f['Ficha Pessoal']

print("--- ROWS 8 TO 15, COLS A TO P (1 to 16) ---")
for r in range(8, 16):
    row_str = []
    for c in range(1, 17):
        v_v = sheet_v.cell(row=r, column=c).value
        v_f = sheet_f.cell(row=r, column=c).value
        if v_v is not None or v_f is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_str.append(f"{col_letter}{r}: v={v_v}/f={v_f}")
    if row_str:
        print(f"Row {r}: {', '.join(row_str)}")
