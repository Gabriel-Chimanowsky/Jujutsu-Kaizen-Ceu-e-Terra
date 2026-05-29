import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

print("=== Searching Excel for 'Estamina' ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(1, 100):
        for col in range(1, 100):
            val = ws.cell(row=row, column=col).value
            if val is not None and "estamina" in str(val).lower():
                col_let = openpyxl.utils.get_column_letter(col)
                print(f"Sheet '{sheet_name}', Cell {col_let}{row}: Value='{val}'")
