import openpyxl

wb = openpyxl.load_workbook('exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx', data_only=True)
sheet = wb['Ficha Pessoal']

# Let's inspect rows 17 to 21
for r in range(17, 22):
    for c in range(15, 25):
        cell_name = openpyxl.utils.get_column_letter(c) + str(r)
        cell = sheet.cell(row=r, column=c)
        print(f"{cell_name} (Val: {cell.value}, Type: {type(cell.value).__name__})")
