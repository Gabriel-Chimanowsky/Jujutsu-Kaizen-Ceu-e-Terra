import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

ws_f = wb_formula['Ficha Pessoal']
ws_v = wb_value['Ficha Pessoal']

print("=== Cell by Cell Scan of PE Area ===")
for r in range(7, 11):
    for c in range(30, 42): # AD (30) to AP (41)
        col_let = openpyxl.utils.get_column_letter(c)
        vf = ws_f.cell(row=r, column=c).value
        vv = ws_v.cell(row=r, column=c).value
        print(f"Cell {col_let}{r}: Formula='{vf}' | Value='{vv}'")
