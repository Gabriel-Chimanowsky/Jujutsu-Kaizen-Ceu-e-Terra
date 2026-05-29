import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

ws_f = wb_formula['Ficha Pessoal']
ws_v = wb_value['Ficha Pessoal']

# Let's inspect rows 8 to 16, columns A to N
print("=== Ficha Pessoal Attributes & Modifiers ===")
for row in range(8, 17):
    row_content = []
    for col in range(1, 16):
        col_let = openpyxl.utils.get_column_letter(col)
        vf = ws_f.cell(row=row, column=col).value
        vv = ws_v.cell(row=row, column=col).value
        if vf is not None or vv is not None:
            row_content.append(f"{col_let}{row}: [F: {vf} | V: {vv}]")
    if row_content:
        print(f"Row {row}:")
        for item in row_content:
            print(f"  {item}")
