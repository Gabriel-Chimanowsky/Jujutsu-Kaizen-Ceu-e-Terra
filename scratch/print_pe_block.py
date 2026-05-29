import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

ws_f = wb_formula['Ficha Pessoal']
ws_v = wb_value['Ficha Pessoal']

print("=== Complete Block Dump of PE Area (Rows 6-12, Cols AC-AP) ===")
for r in range(6, 13):
    row_str = []
    for c in range(29, 43):
        col_let = openpyxl.utils.get_column_letter(c)
        vf = ws_f.cell(row=r, column=c).value
        vv = ws_v.cell(row=r, column=c).value
        row_str.append(f"{col_let}{r}: [F={vf}, V={vv}]")
    print(f"Row {r}:")
    for cell in row_str:
        print(f"  {cell}")
