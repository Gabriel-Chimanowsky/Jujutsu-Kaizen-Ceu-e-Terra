import openpyxl

wb = openpyxl.load_workbook('exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx', data_only=False)
sheet = wb['Ficha Pessoal']

# Let's inspect rows 15, 18, 20
for r in [15, 18, 20]:
    for c in range(15, 24):
        cell_name = openpyxl.utils.get_column_letter(c) + str(r)
        cell = sheet.cell(row=r, column=c)
        if cell.value is not None:
            print(f"{cell_name}: Formula/Val = {cell.value}")
            
# Let's print the merged cells in Ficha Pessoal
print("\n=== MERGED RANGES ===")
for r_range in list(sheet.merged_cells.ranges):
    if "O" in str(r_range) or "P" in str(r_range) or "Q" in str(r_range) or "R" in str(r_range) or "S" in str(r_range) or "T" in str(r_range) or "U" in str(r_range) or "V" in str(r_range):
        print(r_range)
