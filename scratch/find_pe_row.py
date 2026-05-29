import openpyxl

wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
ws_v = wb_value['Ficha Pessoal']

for r in range(5, 16):
    for c in range(28, 45):
        val = ws_v.cell(row=r, column=c).value
        if val is not None:
            col_let = openpyxl.utils.get_column_letter(c)
            print(f"{col_let}{r}: Value={val}")
