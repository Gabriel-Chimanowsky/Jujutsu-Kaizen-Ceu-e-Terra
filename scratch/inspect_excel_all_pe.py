import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
ws_f = wb_formula['Ficha Pessoal']

print("=== Search Ficha Pessoal for references to 'Calculos' ===")
for row in range(1, 100):
    for col in range(1, 100):
        val_f = ws_f.cell(row=row, column=col).value
        if val_f is not None and isinstance(val_f, str) and "=" in val_f:
            if "calculos" in val_f.lower():
                col_let = openpyxl.utils.get_column_letter(col)
                print(f"Cell {col_let}{row}: Formula='{val_f}'")
