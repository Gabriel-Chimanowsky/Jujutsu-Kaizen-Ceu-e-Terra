import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

ws_f = wb_formula['Ficha Pessoal']
ws_v = wb_value['Ficha Pessoal']

def print_area(start_row, end_row, start_col, end_col, title=""):
    print(f"\n=== Area: {title} ===")
    for row in range(start_row, end_row + 1):
        row_content = []
        for col in range(start_col, end_col + 1):
            col_let = openpyxl.utils.get_column_letter(col)
            vf = ws_f.cell(row=row, column=col).value
            vv = ws_v.cell(row=row, column=col).value
            if vf is not None or vv is not None:
                row_content.append(f"{col_let}{row}: [F: {vf} | V: {vv}]")
        if row_content:
            print(f"Row {row}:")
            for item in row_content:
                print(f"  {item}")

# XP and level: around AP4, BD2
print_area(1, 5, 38, 60, "Level & XP Area")

# Pontos de Vida: around O6
print_area(5, 12, 14, 28, "Pontos de Vida Area")

# Pontos de Energia: around AC6
print_area(5, 12, 28, 45, "Pontos de Energia Area")

# Defesa: around AZ6
print_area(5, 12, 48, 60, "Defesa Area")
