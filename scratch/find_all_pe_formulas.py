import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)

for sheet_name in wb_formula.sheetnames:
    ws = wb_formula[sheet_name]
    print(f"\nSearching sheet '{sheet_name}'...")
    for row in range(1, 100):
        for col in range(1, 100):
            val = ws.cell(row=row, column=col).value
            if val is not None and isinstance(val, str) and "=" in val:
                val_l = val.lower()
                # If formula contains BD3 (Level), Con, Sab, Pre, Int, Des, For or a class name
                if any(x in val_l for x in ["bd3", "con", "sab", "pre", "int", "des", "for", "lutador", "controlador", "suporte", "técnica", "tecnica", "combate"]):
                    col_let = openpyxl.utils.get_column_letter(col)
                    print(f"  Cell {col_let}{row}: Formula='{val}'")
