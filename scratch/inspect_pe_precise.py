import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

ws_f = wb_formula['Ficha Pessoal']
ws_v = wb_value['Ficha Pessoal']

print("=== Direct Cell Check in Row 8 (Columns AC-AL) ===")
for col in range(29, 39):
    col_let = openpyxl.utils.get_column_letter(col)
    val_f = ws_f.cell(row=8, column=col).value
    val_v = ws_v.cell(row=8, column=col).value
    print(f"Cell {col_let}8: Formula='{val_f}' | Value='{val_v}'")

print("\n=== Direct Cell Check in Row 9 (Columns AC-AL) ===")
for col in range(29, 39):
    col_let = openpyxl.utils.get_column_letter(col)
    val_f = ws_f.cell(row=9, column=col).value
    val_v = ws_v.cell(row=9, column=col).value
    print(f"Cell {col_let}9: Formula='{val_f}' | Value='{val_v}'")
