import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
if "Calculos" in wb.sheetnames:
    sheet = wb["Calculos"]
    print("--- Sheet Calculos ---")
    for r in range(1, 20):
        row_vals = []
        for c in range(1, 10):
            cell = sheet.cell(row=r, column=c)
            row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {cell.value}")
        print(" | ".join(row_vals))
else:
    print("Sheet Calculos not found. Available sheets:", wb.sheetnames)
