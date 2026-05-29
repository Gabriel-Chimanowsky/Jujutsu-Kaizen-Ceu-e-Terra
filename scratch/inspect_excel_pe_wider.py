import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

ws_f = wb_formula['Ficha Pessoal']
ws_v = wb_value['Ficha Pessoal']

print("=== Ficha Pessoal PE Area (Wider Dump) ===")
for row in range(6, 15):
    for col in range(25, 45): # Y (25) to AS (45)
        vf = ws_f.cell(row=row, column=col).value
        vv = ws_v.cell(row=row, column=col).value
        if vf is not None or vv is not None:
            col_let = openpyxl.utils.get_column_letter(col)
            print(f"  {col_let}{row}: Formula='{vf}' | Value='{vv}'")
