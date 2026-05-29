import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet = wb["Calculos"]
print("--- Column H (Restringido) in Calculos ---")
for r in range(1, sheet.max_row + 1):
    val_f = sheet.cell(row=r, column=8).value
    if val_f is not None:
        print(f"H{r}: {val_f}")
