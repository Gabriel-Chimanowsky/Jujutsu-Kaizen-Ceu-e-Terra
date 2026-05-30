import openpyxl

wb = openpyxl.load_workbook("exemplo/Ficha F&M Shannon 2.5.2.xlsx", data_only=True)
print("Sheet names:", wb.sheetnames)

# Let's inspect some cells in the first few sheets that could have character name
for name in wb.sheetnames:
    ws = wb[name]
    # Search first 10 rows and 10 columns for 'Nome' or similar
    for r in range(1, 15):
        for c in range(1, 15):
            val = ws.cell(row=r, column=c).value
            if val is not None and ('nome' in str(val).lower() or 'shannon' in str(val).lower() or 'yutsuki' in str(val).lower()):
                print(f"[{name}] {openpyxl.utils.get_column_letter(c)}{r}: {val} -> Next cell: {ws.cell(row=r, column=c+1).value}")
