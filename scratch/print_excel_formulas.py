import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
wb_value = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

def print_sheet_details(sheet_name):
    ws_f = wb_formula[sheet_name]
    ws_v = wb_value[sheet_name]
    print(f"\n================ SHEET: {sheet_name} ================")
    for row in range(1, 100):
        row_content = []
        has_val = False
        for col in range(1, 50):
            val_f = ws_f.cell(row=row, column=col).value
            val_v = ws_v.cell(row=row, column=col).value
            if val_f is not None:
                has_val = True
                col_letter = openpyxl.utils.get_column_letter(col)
                row_content.append(f"{col_letter}{row}: [Formula: {val_f} | Value: {val_v}]")
        if has_val:
            print(f"Row {row}:")
            for item in row_content:
                print(f"  {item}")

# Let's inspect 'Calculos' sheet
print_sheet_details("Calculos")
