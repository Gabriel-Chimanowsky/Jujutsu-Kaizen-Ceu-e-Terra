import openpyxl

wb_formula = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
ws_f = wb_formula['Ficha Pessoal']

print("=== Merged ranges for H17 ===")
for r in ws_f.merged_cells.ranges:
    if r.min_row <= 17 <= r.max_row and r.min_col <= 8 <= r.max_col:
        print(f"H17 merged range: {r} | Top-Left cell: {r.start_cell}")
        print("Formula of top-left:", ws_f.cell(row=r.min_row, column=r.min_col).value)
