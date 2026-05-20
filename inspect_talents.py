import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet = wb['Ficha Pessoal']

print("--- TALENTS SECTION (Cols BH-BS, Rows 6-25) ---")
for r in range(6, 25):
    row_vals = []
    for c in range(60, 71): # BH(60) to BT(72)
        val = sheet.cell(row=r, column=c).value
        coord = sheet.cell(row=r, column=c).coordinate
        if val is not None:
            row_vals.append(f"{coord}:{val}")
    if row_vals:
        print(f"Row {r}: {row_vals}")
