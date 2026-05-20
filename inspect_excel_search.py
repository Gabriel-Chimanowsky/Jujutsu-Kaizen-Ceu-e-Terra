import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)

print("--- EXCEL SEARCH ---")
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\nSearching Sheet: {sheetname} ({sheet.max_row} rows, {sheet.max_column} cols)")
    
    # We will search the first 150 rows and first 40 columns
    max_r = min(sheet.max_row, 150)
    max_c = min(sheet.max_column, 40)
    
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                val_str = str(val).strip()
                # If it's a string, check if it contains interesting keywords
                keywords = ["arma", "ataque", "dano", "item", "peso", "qtd", "equipamento", "feitiço", "habilidade", "talento", "mochila", "inventário", "treinamento", "vontade", "fortitude", "reflexos", "astúcia", "integridade", "defesa", "iniciativa"]
                for kw in keywords:
                    if kw in val_str.lower():
                        # print coordinate, cell value
                        coord = sheet.cell(row=r, column=c).coordinate
                        print(f"[{coord}] = '{val_str}'")
                        break
