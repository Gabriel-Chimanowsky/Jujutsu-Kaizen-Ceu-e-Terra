import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet = wb['Perfil Amaldiçoado']

print("--- SPELLS LAYOUT ---")
# Col AA to BK is 27 to 64
for r in range(7, 28):
    row_vals = []
    for c in range(27, 64):
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            coord = sheet.cell(row=r, column=c).coordinate
            row_vals.append(f"{coord}:{val}")
    if row_vals:
        print(f"Row {r}: {row_vals}")
