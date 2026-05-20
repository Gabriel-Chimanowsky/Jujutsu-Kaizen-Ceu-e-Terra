import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet = wb['Treinamentos']

print("--- TREINAMENTOS COLUMNS ---")
# Let's print values in row 4 and row 16
for c in range(1, 60):
    val4 = sheet.cell(row=4, column=c).value
    val16 = sheet.cell(row=16, column=c).value
    if val4:
        print(f"Row 4 Col {c} ({sheet.cell(row=4, column=c).coordinate}): {val4}")
    if val16:
        print(f"Row 16 Col {c} ({sheet.cell(row=16, column=c).coordinate}): {val16}")
