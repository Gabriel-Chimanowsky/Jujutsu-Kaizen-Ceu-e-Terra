import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheetname = wb.sheetnames[4]
print("Accessing Sheet:", repr(sheetname))
sheet = wb[sheetname]

print("--- SUMMONS FIRST BOX (Cols B-V, Rows 1-16) ---")
for r in range(1, 17):
    row_vals = []
    for c in range(2, 23): # B(2) to V(22)
        val = sheet.cell(row=r, column=c).value
        coord = sheet.cell(row=r, column=c).coordinate
        if val is not None:
            row_vals.append(f"{coord}:{val}")
    if row_vals:
        print(f"Row {r}: {row_vals}")
