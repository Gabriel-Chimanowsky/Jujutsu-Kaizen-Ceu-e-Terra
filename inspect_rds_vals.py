import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet = wb['Ficha Pessoal']

print("--- RDs DETAILS ---")
for r in range(21, 25):
    row_vals = []
    for c in range(27, 39): # AA(27) to AL(38)
        val = sheet.cell(row=r, column=c).value
        coord = sheet.cell(row=r, column=c).coordinate
        if val is not None or True: # print everything in this range
            row_vals.append(f"{coord}:{val}")
    print(f"Row {r}: {row_vals}")
