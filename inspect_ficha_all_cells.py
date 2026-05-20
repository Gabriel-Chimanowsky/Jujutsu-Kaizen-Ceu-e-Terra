import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet = wb['Ficha Pessoal']

print("--- ALL NON-EMPTY CELLS (Rows 1-20) ---")
for r in range(1, 21):
    row_vals = []
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            coord = sheet.cell(row=r, column=c).coordinate
            row_vals.append(f"{coord}:{val}")
    if row_vals:
        print(f"Row {r}: {row_vals}")
