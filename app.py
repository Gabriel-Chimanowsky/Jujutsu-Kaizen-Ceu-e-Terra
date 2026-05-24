import os
import sys
import time
import random
import string

# Força inclusão dos diretórios de pacotes locais no Hostinger
base_dir = os.path.dirname(os.path.abspath(__file__))
local_site = os.path.join(base_dir, 'site-packages')

# Sempre inclui no path para garantir, mesmo se os.path.exists falhar por questões de permissão
if local_site not in sys.path:
    sys.path.insert(0, local_site)

# Coleta diagnósticos detalhados do site-packages para ajudar no troubleshooting do Hostinger
exists = os.path.exists(local_site)
is_dir = os.path.isdir(local_site) if exists else False
readable = os.access(local_site, os.R_OK) if exists else False
executable = os.access(local_site, os.X_OK) if exists else False
try:
    stat_val = oct(os.stat(local_site).st_mode & 0o777) if exists else "N/A"
except Exception as e:
    stat_val = f"Error: {str(e)}"

sys.stderr.write(f"[PYTHON DEBUG] Version: {sys.version}\n")
sys.stderr.write(f"[PYTHON DEBUG] Executable: {sys.executable}\n")
sys.stderr.write(f"[PYTHON DEBUG] local_site: {local_site} (Exists: {exists}, IsDir: {is_dir}, Read: {readable}, Exec/Traverse: {executable}, Perms: {stat_val})\n")

user_home = os.path.expanduser('~')
if not user_home or user_home == '/':
    user_home = '/home/u325294696'
for py_ver in ['3.6', '3.7', '3.8', '3.9', '3.10', '3.11']:
    site_path = os.path.join(user_home, '.local', 'lib', f'python{py_ver}', 'site-packages')
    if os.path.exists(site_path) and site_path not in sys.path:
        sys.path.insert(0, site_path)

sys.stderr.write(f"[PYTHON DEBUG] Path: {sys.path}\n")
sys.stderr.flush()

# Habilita suporte a MySQL nativo via PyMySQL
try:
    import pymysql
    pymysql.install_as_MySQLdb()
except ImportError:
    pass

from flask import Flask, request, jsonify, render_template, redirect, url_for, flash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from models import db, User, Character, Status, Attributes, Lobby, xp_to_level, XP_LEVELS
import json

app = Flask(__name__)
# Use environment variable for production, fallback for dev
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'chave_secreta_jujutsu_rpg_super_segura_dev')
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

@app.after_request
def add_header(r):
    # Força desativação de cache para que mudanças no HTML/JS do Vite sejam refletidas imediatamente
    r.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    r.headers["Pragma"] = "no-cache"
    r.headers["Expires"] = "0"
    return r

base_dir = os.path.abspath(os.path.dirname(__file__))
_default_db = 'sqlite:///' + os.path.join(base_dir, 'database.db')
db_url = os.environ.get('DATABASE_URL', _default_db)
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
elif db_url.startswith('mysql://'):
    db_url = db_url.replace('mysql://', 'mysql+pymysql://', 1)

if db_url.startswith('mysql+pymysql://'):
    if ('localhost' in db_url or '127.0.0.1' in db_url) and 'unix_socket' not in db_url:
        possible_sockets = [
            '/var/lib/mysql/mysql.sock',
            '/var/run/mysqld/mysqld.sock',
            '/tmp/mysql.sock'
        ]
        for socket_path in possible_sockets:
            if os.path.exists(socket_path):
                separator = '&' if '?' in db_url else '?'
                db_url += f"{separator}unix_socket={socket_path}"
                sys.stderr.write(f"[PYTHON DEBUG] Unix Socket MySQL detectado e ativado: {socket_path}\n")
                sys.stderr.flush()
                break

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(base_dir, 'static', 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def _gen_lobby_code(length=6):
    """Gera um código único de lobby (e.g. 'KUROI1')."""
    chars = string.ascii_uppercase + string.digits
    while True:
        code = ''.join(random.choices(chars, k=length))
        if not Lobby.query.filter_by(codigo=code).first():
            return code

db.init_app(app)

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Create tables + run migrations before first request
with app.app_context():
    db.create_all()
    
    # Auto-seed: se nao houver nenhum usuario, cria o usuario mestre padrao
    try:
        mestre = User.query.filter_by(username='mestre').first()
        if not mestre:
            mestre = User(username='mestre', role='Mestre')
            mestre.set_password('mestre123')
            db.session.add(mestre)
            db.session.commit()
            print("[INFO] Banco de dados auto-semeado: usuario 'mestre' / senha 'mestre123' criado.")
        elif mestre.password_hash.startswith('scrypt:'):
            # Se o hash for scrypt (gerado no PC local com Python 3.10+), recria usando pbkdf2:sha256 para Python 3.6
            print("[INFO] Corrigindo hash do mestre de scrypt para pbkdf2:sha256...")
            mestre.set_password('mestre123')
            db.session.commit()
            print("[INFO] Hash do mestre corrigido com sucesso.")
    except Exception as e:
        print("Erro ao semear/corrigir banco de dados:", e)

    # Migrações automáticas genéricas (compatível com SQLite, MySQL e outros bancos)
    try:
        from sqlalchemy import inspect, text
        inspector = inspect(db.engine)
        
        # ── characters table migrations ──
        if 'characters' in inspector.get_table_names():
            columns = [c['name'] for c in inspector.get_columns('characters')]
            for col, col_type in {
                'resistencias': "TEXT",
                'rds': "TEXT",
                'habilidades_talentos': "TEXT",
                'dados_vida': "TEXT",
                'anotacoes': "TEXT",
                'caracteristicas': "TEXT",
                'configuracoes': "TEXT",
                'dominio': "TEXT",
                'recent_logs': "TEXT",
                'pontos_atributos': "INTEGER",
                'peso': "TEXT",
                'altura': "TEXT",
                'afiliacao': "TEXT",
                'votos_ativos': "TEXT"
            }.items():
                if col not in columns:
                    sql_type = col_type
                    # SQLite exige valores padrão específicos para ADD COLUMN se as colunas forem NOT NULL ou por consistência
                    if db.engine.name == 'sqlite':
                        if col in ['resistencias', 'rds', 'configuracoes', 'dominio']:
                            sql_type += " DEFAULT '{}'"
                        elif col in ['habilidades_talentos', 'caracteristicas', 'recent_logs']:
                            sql_type += " DEFAULT '[]'"
                        elif col == 'anotacoes':
                            sql_type += " DEFAULT ''"
                        elif col == 'pontos_atributos':
                            sql_type += " DEFAULT 0"
                        elif col == 'peso':
                            sql_type += " DEFAULT '72kg'"
                        elif col == 'altura':
                            sql_type += " DEFAULT '1.82m'"
                        elif col == 'afiliacao':
                            sql_type += " DEFAULT 'Colégio Técnico de Jujutsu'"
                        elif col == 'votos_ativos':
                            sql_type += " DEFAULT 'Revelação da Técnica (+2 CD Feitiços)'"
                    
                    db.session.execute(text(f"ALTER TABLE characters ADD COLUMN {col} {sql_type}"))
            db.session.commit()

        # ── users table migrations ──
        if 'users' in inspector.get_table_names():
            columns = [c['name'] for c in inspector.get_columns('users')]
            if 'lobby_id' not in columns:
                db.session.execute(text("ALTER TABLE users ADD COLUMN lobby_id INTEGER REFERENCES lobbies(id)"))
                db.session.commit()
            
            # Retroactive migration for user_lobbies sintonization history
            try:
                users_with_lobby = User.query.filter(User.lobby_id.isnot(None)).all()
                for u in users_with_lobby:
                    lobby_obj = Lobby.query.get(u.lobby_id)
                    if lobby_obj and lobby_obj.ativo and lobby_obj not in u.joined_lobbies:
                        u.joined_lobbies.append(lobby_obj)
                db.session.commit()
            except Exception as ex:
                print("Erro ao migrar sintonização retroativa user_lobbies:", ex)

        # ── lobbies table migrations ──
        if 'lobbies' in inspector.get_table_names():
            columns = [c['name'] for c in inspector.get_columns('lobbies')]
            if 'vtt_state' not in columns:
                db.session.execute(text("ALTER TABLE lobbies ADD COLUMN vtt_state TEXT"))
                db.session.commit()
    except Exception as e:
        print("Erro durante a migracao automatica genérica:", e)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/auth/status')
def api_auth_status():
    if current_user.is_authenticated:
        char = Character.query.filter_by(user_id=current_user.id).first()
        return jsonify({
            'authenticated': True,
            'user_id': current_user.id,
            'username': current_user.username,
            'role': current_user.role,
            'character_id': char.id if char else None,
            'character_nome': char.nome if char else None,
            'lobby_id': current_user.lobby_id
        })
    return jsonify({'authenticated': False})

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if request.headers.get('Accept') == 'application/json' or request.is_json:
            return jsonify({'message': 'Logged in successfully', 'redirect': url_for('lobby')})
        return redirect(url_for('lobby'))
        
    if request.method == 'POST':
        # Handles both JSON and Form data for flexibility
        data = request.get_json() if request.is_json else request.form
        username = data.get('username')
        password = data.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            if request.is_json:
                return jsonify({'message': 'Logged in successfully', 'redirect': url_for('lobby')})
            return redirect(url_for('lobby'))
            
        if request.is_json:
            return jsonify({'error': 'Invalid username or password'}), 401
        flash('Invalid username or password')
        
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        if request.headers.get('Accept') == 'application/json' or request.is_json:
            return jsonify({'message': 'Already logged in', 'redirect': url_for('lobby')})
        return redirect(url_for('lobby'))
        
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username')
        password = data.get('password')
        role = data.get('role', 'Jogador') # 'Mestre' or 'Jogador'
        
        if User.query.filter_by(username=username).first():
            if request.is_json:
                return jsonify({'error': 'Username already exists'}), 400
            flash('Username already exists')
            return redirect(url_for('register'))
            
        new_user = User(username=username, role=role)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        
        if request.is_json:
            return jsonify({'message': 'User registered successfully'}), 201
        flash('Registration successful. Please log in.')
        return redirect(url_for('login'))
        
    return render_template('index.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ═══════════════════════════════════════════════════════════════════
# LOBBY MANAGEMENT ROUTES
# ═══════════════════════════════════════════════════════════════════

@app.route('/lobby/criar', methods=['POST'])
@login_required
def lobby_criar():
    """Mestre cria um novo lobby."""
    if current_user.role != 'Mestre':
        return jsonify({'error': 'Apenas Mestres podem criar lobbies.'}), 403
    data = request.get_json() or {}
    nome = data.get('nome', 'Domínio do Mestre').strip() or 'Domínio do Mestre'
    # Remove mestre do lobby anterior se houver
    current_user.lobby_id = None
    db.session.flush()
    codigo = _gen_lobby_code()
    lobby = Lobby(nome=nome, codigo=codigo, master_id=current_user.id)
    db.session.add(lobby)
    db.session.flush()  # get lobby.id
    current_user.lobby_id = lobby.id
    db.session.commit()
    return jsonify({'ok': True, 'lobby': lobby.to_dict(), 'codigo': codigo})


@app.route('/lobby/entrar', methods=['POST'])
@login_required
def lobby_entrar():
    """Jogador entra em lobby via código."""
    data = request.get_json() or {}
    codigo = (data.get('codigo') or '').strip().upper()
    if not codigo:
        return jsonify({'error': 'Código obrigatório.'}), 400
    lobby = Lobby.query.filter_by(codigo=codigo, ativo=True).first()
    if not lobby:
        return jsonify({'error': 'Lobby não encontrado. Verifique o código.'}), 404
    
    # Adiciona na lista de sintonizados do usuário
    if lobby not in current_user.joined_lobbies:
        current_user.joined_lobbies.append(lobby)
        
    current_user.lobby_id = lobby.id
    db.session.commit()
    return jsonify({'ok': True, 'lobby': lobby.to_dict()})


@app.route('/lobby/select/<int:lobby_id>', methods=['POST'])
@login_required
def lobby_select(lobby_id):
    """Jogador seleciona um lobby da sua lista para entrar."""
    lobby = Lobby.query.filter_by(id=lobby_id, ativo=True).first()
    if not lobby:
        return jsonify({'error': 'Lobby não encontrado ou inativo.'}), 404
    
    # Garante que o mestre ou participantes possam entrar
    if lobby.master_id != current_user.id and lobby not in current_user.joined_lobbies:
        current_user.joined_lobbies.append(lobby)
        
    current_user.lobby_id = lobby.id
    db.session.commit()
    return jsonify({'ok': True, 'lobby': lobby.to_dict()})


@app.route('/lobby/remover/<int:lobby_id>', methods=['POST'])
@login_required
def lobby_remover(lobby_id):
    """Jogador remove um lobby da sua lista (deixa de sintonizar)."""
    lobby = Lobby.query.get(lobby_id)
    if not lobby:
        return jsonify({'error': 'Lobby não encontrado.'}), 404
    if lobby in current_user.joined_lobbies:
        current_user.joined_lobbies.remove(lobby)
    if current_user.lobby_id == lobby.id:
        current_user.lobby_id = None
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/lobby/sair', methods=['POST'])
@login_required
def lobby_sair():
    """Player/Mestre sai da visualização ativa do lobby (mas permanece sintonizado)."""
    current_user.lobby_id = None
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/lobby/kick/<int:user_id>', methods=['POST'])
@login_required
def lobby_kick(user_id):
    """Mestre remove jogador do lobby."""
    if current_user.role != 'Mestre':
        return jsonify({'error': 'Apenas Mestres podem remover jogadores.'}), 403
    if not current_user.lobby_id:
        return jsonify({'error': 'Você não está em nenhum lobby.'}), 400
    lobby = Lobby.query.get(current_user.lobby_id)
    if not lobby or lobby.master_id != current_user.id:
        return jsonify({'error': 'Não autorizado.'}), 403
    target = User.query.get_or_404(user_id)
    if target.lobby_id != current_user.lobby_id:
        return jsonify({'error': 'Jogador não está no seu lobby.'}), 404
    if target.id == current_user.id:
        return jsonify({'error': 'Mestre não pode se remover. Use Fechar Lobby.'}), 400
    target.lobby_id = None
    db.session.commit()
    return jsonify({'ok': True, 'kicked_user': target.username})


@app.route('/lobby/add', methods=['POST'])
@login_required
def lobby_add():
    """Mestre adiciona jogador ao lobby pelo username."""
    if current_user.role != 'Mestre':
        return jsonify({'error': 'Apenas Mestres podem adicionar jogadores.'}), 403
    if not current_user.lobby_id:
        return jsonify({'error': 'Você não está em nenhum lobby.'}), 400
    lobby = Lobby.query.get(current_user.lobby_id)
    if not lobby or lobby.master_id != current_user.id:
        return jsonify({'error': 'Não autorizado.'}), 403
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    target = User.query.filter_by(username=username).first()
    if not target:
        return jsonify({'error': f'Usuário "{username}" não encontrado.'}), 404
    if target.lobby_id is not None and target.lobby_id != current_user.lobby_id:
        return jsonify({'error': f'{username} já está em outro lobby.'}), 409
    target.lobby_id = current_user.lobby_id
    db.session.commit()
    return jsonify({'ok': True, 'added_user': target.username})


@app.route('/lobby/fechar', methods=['POST'])
@login_required
def lobby_fechar():
    """Mestre fecha/exclui o lobby atual."""
    if current_user.role != 'Mestre':
        return jsonify({'error': 'Apenas Mestres podem fechar lobbies.'}), 403
    if not current_user.lobby_id:
        return jsonify({'error': 'Você não está em nenhum lobby.'}), 400
    lobby = Lobby.query.get(current_user.lobby_id)
    if not lobby or lobby.master_id != current_user.id:
        return jsonify({'error': 'Não autorizado.'}), 403
    for member in lobby.membros.all():
        member.lobby_id = None
    lobby.ativo = False
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/lobby_status')
@login_required
def api_lobby_status():
    """Retorna o estado atual do lobby do usuário (para polling JS)."""
    if not current_user.lobby_id:
        return jsonify({'in_lobby': False})
    lobby = Lobby.query.get(current_user.lobby_id)
    if not lobby or not lobby.ativo:
        current_user.lobby_id = None
        db.session.commit()
        return jsonify({'in_lobby': False, 'reason': 'Lobby foi fechado.'})
    membros = []
    for u in lobby.membros.all():
        char = Character.query.filter_by(user_id=u.id).first()
        membros.append({
            'user_id': u.id,
            'username': u.username,
            'is_master': u.id == lobby.master_id,
            'char_id': char.id if char else None,
            'char_nome': char.nome if char else None
        })
    return jsonify({
        'in_lobby': True,
        'lobby': lobby.to_dict(),
        'is_master': current_user.id == lobby.master_id,
        'membros': membros
    })


@app.route('/api/dar_xp/<int:char_id>', methods=['POST'])
@login_required
def dar_xp(char_id):
    """Mestre dá XP a um personagem. Auto-calcula o nível."""
    if current_user.role != 'Mestre':
        return jsonify({'error': 'Apenas Mestres podem dar XP.'}), 403
    char = Character.query.get_or_404(char_id)
    data = request.get_json() or {}
    quantidade = int(data.get('quantidade', 0))
    if quantidade == 0:
        return jsonify({'error': 'Quantidade de XP inválida.'}), 400
    old_level = char.nivel or 1
    char.xp = max(0, (char.xp or 0) + quantidade)
    new_level = xp_to_level(char.xp)
    level_up = new_level > old_level
    if new_level != old_level:
        old_pv_max = char.status.pv_max if char.status else 10
        old_pe_max = char.status.pe_max if char.status else 0
        
        char.nivel = new_level
        
        if level_up:
            level_diff = new_level - old_level
            char.pontos_atributos = (char.pontos_atributos or 0) + (level_diff * 2)
            
        if char.status:
            scale_current_status_proportionally(char, old_pv_max, old_pe_max)
            
    db.session.commit()
    # Log the XP grant
    try:
        logs = json.loads(char.recent_logs or '[]')
        sign = '+' if quantidade >= 0 else ''
        msg = f'[XP] {sign}{quantidade} XP recebido'
        if level_up:
            msg += f'  -> LEVEL UP! Nível {new_level}!'
        logs.insert(0, {
            'type': 'xp',
            'title': '[Mestre] XP Concedido',
            'content': msg,
            'timestamp': time.strftime('%H:%M')
        })
        char.recent_logs = json.dumps(logs[:20])
        db.session.commit()
    except:
        pass
    return jsonify({
        'ok': True,
        'xp_total': char.xp,
        'nivel': char.nivel,
        'level_up': level_up,
        'old_level': old_level,
        'new_level': new_level,
        'char_nome': char.nome
    })


# ═══════════════════════════════════════════════════════════════════
# LOBBY VIEW
# ═══════════════════════════════════════════════════════════════════

@app.route('/lobby', methods=['GET'])
@login_required
def lobby():
    # If user is not in a lobby, show entry screen with sintonized lobbies
    if not current_user.lobby_id:
        if request.headers.get('Accept') == 'application/json' or request.args.get('format') == 'json' or request.is_json:
            joined = [l.to_dict() for l in current_user.joined_lobbies if l.ativo]
            if current_user.role == 'Mestre':
                for l in current_user.lobbies_criados:
                    if l.ativo and l.id not in [jl['id'] for jl in joined]:
                        joined.append(l.to_dict())
            return jsonify({'in_lobby': False, 'joined_lobbies': joined})
        return render_template('index.html')

    lobby_obj = Lobby.query.get(current_user.lobby_id)
    if not lobby_obj or not lobby_obj.ativo:
        current_user.lobby_id = None
        db.session.commit()
        if request.headers.get('Accept') == 'application/json' or request.args.get('format') == 'json' or request.is_json:
            joined = [l.to_dict() for l in current_user.joined_lobbies if l.ativo]
            if current_user.role == 'Mestre':
                for l in current_user.lobbies_criados:
                    if l.ativo and l.id not in [jl['id'] for jl in joined]:
                        joined.append(l.to_dict())
            return jsonify({'in_lobby': False, 'reason': 'Lobby foi fechado.', 'joined_lobbies': joined})
        return render_template('index.html')

    # Show only characters belonging to users in this lobby
    lobby_user_ids = [u.id for u in lobby_obj.membros.all()]
    characters_objs = Character.query.filter(Character.user_id.in_(lobby_user_ids)).all()
    char_data = []
    for char in characters_objs:
        def _mod(v): return (v - 10) // 2
        attrs = char.attributes
        forca        = attrs.forca        if attrs else 10
        destreza     = attrs.destreza     if attrs else 10
        constituicao = attrs.constituicao if attrs else 10
        inteligencia = attrs.inteligencia if attrs else 10
        sabedoria    = attrs.sabedoria    if attrs else 10
        presenca     = attrs.presenca     if attrs else 10
        char_data.append({
            'id': char.id,
            'nome': char.nome,
            'imagem_url': char.imagem_url,
            'grau': char.grau,
            'nivel': char.nivel,
            'xp': char.xp or 0,
            'especializacao': char.especializacao,
            'origem': char.origem,
            'user_id': char.user_id,
            'pv_atual':          char.status.pv_atual          if char.status else 0,
            'pv_max':            char.status.pv_max             if char.status else 0,
            'pe_atual':          char.status.pe_atual           if char.status else 0,
            'pe_max':            char.status.pe_max             if char.status else 0,
            'integridade_atual': char.status.integridade_atual  if char.status else 0,
            'integridade_max':   char.status.integridade_max    if char.status else 0,
            'estado_alma':       char.status.estado_alma        if char.status else 'Desconhecido',
            'cor_energia':       char.cor_energia,
            'ataques':           json.loads(char.ataques           or '[]'),
            'feiticos':          json.loads(char.feiticos          or '[]'),
            'habilidades_talentos': json.loads(char.habilidades_talentos or '[]'),
            'recent_logs':       json.loads(char.recent_logs       or '[]'),
            'inventario':        json.loads(char.inventario        or '[]'),
            'attributes': {
                'forca': forca, 'destreza': destreza, 'constituicao': constituicao,
                'inteligencia': inteligencia, 'sabedoria': sabedoria, 'presenca': presenca
            },
            'mods': {
                'forca': _mod(forca), 'destreza': _mod(destreza), 'constituicao': _mod(constituicao),
                'inteligencia': _mod(inteligencia), 'sabedoria': _mod(sabedoria), 'presenca': _mod(presenca)
            }
        })

    is_master = (current_user.id == lobby_obj.master_id)
    members_list = []
    for u in lobby_obj.membros.all():
        ch = Character.query.filter_by(user_id=u.id).first()
        members_list.append({
            'user_id':   u.id,
            'username':  u.username,
            'is_master': u.id == lobby_obj.master_id,
            'char_id':   ch.id    if ch else None,
            'char_nome': ch.nome  if ch else '—',
            'char_nivel': ch.nivel if ch else 1,
            'char_xp':   ch.xp    if ch else 0,
        })

    if request.headers.get('Accept') == 'application/json' or request.args.get('format') == 'json' or request.is_json:
        return jsonify({
            'in_lobby': True,
            'characters': char_data,
            'lobby': lobby_obj.to_dict(),
            'is_master': is_master,
            'members': members_list,
            'current_user_id': current_user.id,
            'lobby_codigo': lobby_obj.codigo,
            'vtt_state': json.loads(lobby_obj.vtt_state) if lobby_obj.vtt_state else None
        })

    return render_template('index.html')



@app.route('/lobby/vtt/update', methods=['POST'])
@login_required
def lobby_vtt_update():
    if not current_user.lobby_id:
        return jsonify({'error': 'Você não está em um lobby!'}), 400
    
    lobby_obj = Lobby.query.get(current_user.lobby_id)
    if not lobby_obj or not lobby_obj.ativo:
        return jsonify({'error': 'Lobby inativo ou inexistente!'}), 400
    
    try:
        vtt_data = request.json
        lobby_obj.vtt_state = json.dumps(vtt_data)
        db.session.commit()
        return jsonify({'ok': True, 'vtt_state': vtt_data})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Falha ao atualizar VTT: {str(e)}'}), 500

@app.route('/create_character', methods=['GET', 'POST'])
@login_required
def create_character():
    if request.method == 'POST':
        nome = request.form.get('nome')
        origem = request.form.get('origem')
        especializacao = request.form.get('especializacao')
        peso = request.form.get('peso', '72kg')
        altura = request.form.get('altura', '1.82m')
        afiliacao = request.form.get('afiliacao', 'Colégio Técnico de Jujutsu')
        votos_ativos = request.form.get('votos_ativos', 'Revelação da Técnica (+2 CD Feitiços)')
        
        new_char = Character(
            user_id=current_user.id,
            nome=nome,
            origem=origem,
            especializacao=especializacao,
            peso=peso,
            altura=altura,
            afiliacao=afiliacao,
            votos_ativos=votos_ativos
        )
        db.session.add(new_char)
        db.session.commit()
        
        new_status = Status(character_id=new_char.id)
        new_attr = Attributes(
            character_id=new_char.id,
            forca=int(request.form.get('forca', 10)),
            destreza=int(request.form.get('destreza', 10)),
            constituicao=int(request.form.get('constituicao', 10)),
            inteligencia=int(request.form.get('inteligencia', 10)),
            sabedoria=int(request.form.get('sabedoria', 10)),
            presenca=int(request.form.get('presenca', 10))
        )
        db.session.add(new_status)
        db.session.add(new_attr)
        db.session.commit()
        
        # Heals character to math PV/PE computed properties after attributes are set
        new_status.pv_atual = new_status.pv_max
        new_status.pe_atual = new_status.pe_max
        new_status.integridade_atual = new_status.integridade_max
        db.session.commit()
        
        return redirect(url_for('lobby'))
        
    return render_template('index.html')

@app.route('/ficha/<int:character_id>', methods=['GET'])
@login_required
def ficha(character_id):
    char = Character.query.get_or_404(character_id)
    
    # Jogador can only see their own character
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        if request.headers.get('Accept') == 'application/json' or request.args.get('format') == 'json':
            return jsonify({'error': 'Unauthorized'}), 403
        flash('Você não tem permissão para acessar esta ficha.')
        return redirect(url_for('lobby'))
        
    char_data = {
        'id': char.id,
        'nome': char.nome,
        'origem': char.origem,
        'especializacao': char.especializacao,
        'grau': char.grau,
        'nivel': char.nivel,
        'xp': char.xp,
        'imagem_url': char.imagem_url,
        'cor_energia': char.cor_energia,
        'resistencias': json.loads(char.resistencias or '{}'),
        'rds': json.loads(char.rds or '{}'),
        'habilidades_talentos': json.loads(char.habilidades_talentos or '[]'),
        'dados_vida': json.loads(char.dados_vida or '{}'),
        'anotacoes': char.anotacoes or '',
        'caracteristicas': json.loads(char.caracteristicas or '[]'),
        'configuracoes': json.loads(char.configuracoes or '{}'),
        'iniciativa': char.iniciativa,
        'atencao_passiva': char.atencao_passiva,
        'cd_especializacao': char.cd_especializacao,
        'training_bonus': char.training_bonus,
        'half_level': char.half_level,
        'status': {
            'pv_atual': char.status.pv_atual if char.status else 0,
            'pv_max': char.status.pv_max if char.status else 0,
            'pe_atual': char.status.pe_atual if char.status else 0,
            'pe_max': char.status.pe_max if char.status else 0,
            'integridade_atual': char.status.integridade_atual if char.status else 0,
            'integridade_max': char.status.integridade_max if char.status else 0,
            'estado_alma': char.status.estado_alma if char.status else 'Desconhecido',
            'falhas_morte': char.status.falhas_morte if char.status else 0,
            'sucessos_morte': char.status.sucessos_morte if char.status else 0,
            'defesa': char.status.defesa if char.status else 0
        },
        'attributes': {
            'forca': char.attributes.forca if char.attributes else 0,
            'destreza': char.attributes.destreza if char.attributes else 0,
            'constituicao': char.attributes.constituicao if char.attributes else 0,
            'inteligencia': char.attributes.inteligencia if char.attributes else 0,
            'sabedoria': char.attributes.sabedoria if char.attributes else 0,
            'presenca': char.attributes.presenca if char.attributes else 0
        }
    }
    
    if request.headers.get('Accept') == 'application/json' or request.args.get('format') == 'json':
        return jsonify(get_character_json(char))
        
    return render_template('index.html')

@app.route('/legacy-ficha/<int:character_id>', methods=['GET'])
@login_required
def legacy_ficha(character_id):
    char = Character.query.get_or_404(character_id)
    
    # Jogador can only see their own character
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        flash('Você não tem permissão para acessar esta ficha.')
        return redirect(url_for('lobby'))
        
    return render_template('ficha.html', character=char)

@app.route('/api/update_status/<int:character_id>', methods=['POST'])
@login_required
def update_status(character_id):
    char = Character.query.get_or_404(character_id)
    
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    if not data or not char.status:
        return jsonify({'error': 'Invalid data or character has no status initialized'}), 400
        
    if 'pv_delta' in data:
        char.status.pv_atual = max(0, min(char.status.pv_max, char.status.pv_atual + data['pv_delta']))
    
    if 'pe_delta' in data:
        char.status.pe_atual = max(0, min(char.status.pe_max, char.status.pe_atual + data['pe_delta']))
        
    if 'integridade_delta' in data:
        char.status.integridade_atual = max(0, min(char.status.integridade_max, char.status.integridade_atual + data['integridade_delta']))
        
    if 'falhas_delta' in data:
        char.status.falhas_morte += data['falhas_delta']
        
    if 'sucessos_delta' in data:
        char.status.sucessos_morte += data['sucessos_delta']
        
    db.session.commit()
    
    return jsonify({
        'message': 'Status updated successfully',
        'character': get_character_json(char)
    })

@app.route('/api/update_energy_color/<int:character_id>', methods=['POST'])
@login_required
def update_energy_color(character_id):
    if current_user.role != 'Mestre':
        return jsonify({'error': 'Unauthorized. Only Mestre can change energy color.'}), 403
        
    char = Character.query.get_or_404(character_id)
    data = request.get_json()
    
    if not data or 'cor_energia' not in data:
        return jsonify({'error': 'Invalid data. Please provide cor_energia.'}), 400
        
    char.cor_energia = data['cor_energia']
    db.session.commit()
    
    return jsonify({
        'message': 'Energy color updated successfully',
        'cor_energia': char.cor_energia
    })

@app.route('/api/update_attributes/<int:character_id>', methods=['POST'])
@login_required
def update_attributes(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role != 'Mestre':
        return jsonify({'error': 'Apenas o Mestre pode alterar atributos livremente.'}), 403
        
    data = request.get_json()
    if not data or 'attr' not in data or 'delta' not in data:
        return jsonify({'error': 'Invalid data'}), 400
        
    attr_map = {
        'forca': 'forca',
        'destreza': 'destreza',
        'constituicao': 'constituicao',
        'inteligencia': 'inteligencia',
        'sabedoria': 'sabedoria',
        'presenca': 'presenca'
    }
    
    attr = data['attr']
    if attr in attr_map:
        field = attr_map[attr]
        
        old_pv_max = char.status.pv_max if char.status else 10
        old_pe_max = char.status.pe_max if char.status else 0
        
        current_val = getattr(char.attributes, field)
        setattr(char.attributes, field, max(1, current_val + data['delta']))
        
        if char.status:
            scale_current_status_proportionally(char, old_pv_max, old_pe_max)
            
        db.session.commit()
        return jsonify({
            'message': 'Attribute updated', 
            'new_val': getattr(char.attributes, field), 
            'attr': attr,
            'character': get_character_json(char)
        })
        
    return jsonify({'error': 'Invalid attribute'}), 400

@app.route('/api/confirm_attributes/<int:character_id>', methods=['POST'])
@login_required
def confirm_attributes(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Você não tem permissão para editar este personagem.'}), 403
        
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Dados inválidos'}), 400
        
    attrs_keys = ['forca', 'destreza', 'constituicao', 'inteligencia', 'sabedoria', 'presenca']
    proposed = {}
    for key in attrs_keys:
        if key not in data:
            return jsonify({'error': f'Atributo "{key}" ausente nos dados de envio.'}), 400
        try:
            proposed[key] = int(data[key])
        except (ValueError, TypeError):
            return jsonify({'error': f'Valor inválido para o atributo "{key}".'}), 400
            
    total_cost = 0
    for key in attrs_keys:
        db_val = getattr(char.attributes, key)
        new_val = proposed[key]
        if new_val < db_val:
            return jsonify({'error': f'Não é permitido diminuir atributos após confirmar (Atributo {key}: banco={db_val}, enviado={new_val}).'}), 400
        total_cost += (new_val - db_val)
        
    if total_cost == 0:
        return jsonify({'error': 'Nenhum ponto foi distribuído.'}), 400
        
    if total_cost > (char.pontos_atributos or 0):
        return jsonify({'error': f'Você está tentando gastar {total_cost} pontos, mas só possui {char.pontos_atributos or 0} disponíveis.'}), 400
        
    old_pv_max = char.status.pv_max if char.status else 10
    old_pe_max = char.status.pe_max if char.status else 0
    
    for key in attrs_keys:
        setattr(char.attributes, key, proposed[key])
        
    char.pontos_atributos = max(0, (char.pontos_atributos or 0) - total_cost)
    
    if char.status:
        scale_current_status_proportionally(char, old_pv_max, old_pe_max)
        
    try:
        logs = json.loads(char.recent_logs or '[]')
        msg = f'Evolução confirmada: +{total_cost} pontos distribuídos!'
        logs.insert(0, {
            'type': 'evolution',
            'title': 'Evolução de Atributos',
            'content': msg,
            'timestamp': time.strftime('%H:%M')
        })
        char.recent_logs = json.dumps(logs[:20])
    except Exception as e:
        print("Erro ao adicionar log de evolução:", e)
        
    db.session.commit()
    
    return jsonify({
        'message': 'Evolução realizada com sucesso!',
        'character': get_character_json(char)
    })

@app.route('/api/get_inventory/<int:character_id>')
@login_required
def get_inventory(character_id):
    char = Character.query.get_or_404(character_id)
    return jsonify(json.loads(char.inventario or '[]'))

@app.route('/api/add_item/<int:character_id>', methods=['POST'])
@login_required
def add_item(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    inventory = json.loads(char.inventario or '[]')
    
    new_item = {
        'id': len(inventory) + 1,
        'nome': data.get('nome', 'Item Sem Nome'),
        'qtd': int(data.get('qtd', 1)),
        'peso': float(data.get('peso', 0.0))
    }
    
    inventory.append(new_item)
    char.inventario = json.dumps(inventory)
    db.session.commit()
    
    return jsonify(inventory)

@app.route('/api/delete_item/<int:character_id>/<int:item_id>', methods=['DELETE'])
@login_required
def delete_item(character_id, item_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    inventory = json.loads(char.inventario or '[]')
    inventory = [item for item in inventory if item['id'] != item_id]
    
    char.inventario = json.dumps(inventory)
    db.session.commit()
    
    return jsonify(inventory)

@app.route('/api/update_item/<int:character_id>/<int:item_id>', methods=['POST'])
@login_required
def update_item(character_id, item_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    inventory = json.loads(char.inventario or '[]')
    
    for item in inventory:
        if item['id'] == item_id:
            if 'nome' in data: item['nome'] = data['nome']
            if 'qtd' in data: item['qtd'] = int(data['qtd'])
            if 'peso' in data: item['peso'] = float(data['peso'])
            if 'equipado' in data: item['equipado'] = bool(data['equipado'])
            break
            
    char.inventario = json.dumps(inventory)
    db.session.commit()
    return jsonify(inventory)

@app.route('/api/update_character_basics/<int:character_id>', methods=['POST'])
@login_required
def update_character_basics(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Invalid data'}), 400
        
    recalc_hp_pe = False
    if 'nome' in data:
        char.nome = data['nome']
    if 'nivel' in data and int(data['nivel']) != char.nivel:
        char.nivel = int(data['nivel'])
        recalc_hp_pe = True
    if 'xp' in data:
        char.xp = int(data['xp'])
    if 'grau' in data:
        char.grau = data['grau']
    if 'origem' in data and data['origem'] != char.origem:
        char.origem = data['origem']
        recalc_hp_pe = True
    if 'especializacao' in data and data['especializacao'] != char.especializacao:
        char.especializacao = data['especializacao']
        recalc_hp_pe = True
    if 'cor_energia' in data:
        char.cor_energia = data['cor_energia']
    if 'imagem_url' in data:
        char.imagem_url = data['imagem_url']

    # Campos de Registro Físico
    if 'peso' in data:
        char.peso = data['peso']
    if 'altura' in data:
        char.altura = data['altura']
    if 'afiliacao' in data:
        char.afiliacao = data['afiliacao']
    if 'votos_ativos' in data:
        char.votos_ativos = data['votos_ativos']
        
    # Recalculates HP and PE only if basic stats affecting properties were changed
    if recalc_hp_pe and char.status:
        char.status.pv_atual = char.status.pv_max
        char.status.pe_atual = char.status.pe_max
        char.status.integridade_atual = char.status.integridade_max
        
    db.session.commit()
    
    return jsonify({
        'message': 'Basics updated successfully',
        'character': get_character_json(char)
    })

@app.route('/api/get_pericias/<int:character_id>')
@login_required
def get_pericias(character_id):
    char = Character.query.get_or_404(character_id)
    if not char.pericias or char.pericias in ['{}', '[]']:
        # This will initialize default pericias for this char
        char.__init__(nome=char.nome, user_id=char.user_id)
        db.session.commit()
    return jsonify(json.loads(char.pericias))

@app.route('/api/update_pericias/<int:character_id>', methods=['POST'])
@login_required
def update_pericias(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    if not data or 'nome' not in data:
        return jsonify({'error': 'Invalid data'}), 400
        
    pericias = json.loads(char.pericias or '{}')
    p_name = data['nome']
    
    if p_name in pericias:
        if 'treinada' in data:
            pericias[p_name]['treinada'] = bool(data['treinada'])
        if 'mestre' in data:
            pericias[p_name]['mestre'] = bool(data['mestre'])
        if 'bonus' in data:
            pericias[p_name]['bonus'] = int(data['bonus'])
            
        char.pericias = json.dumps(pericias)
        db.session.commit()
        return jsonify({
            'pericias': pericias,
            'character': get_character_json(char)
        })
        
    return jsonify({'error': 'Skill not found'}), 404

@app.route('/api/get_attacks/<int:character_id>')
@login_required
def get_attacks(character_id):
    char = Character.query.get_or_404(character_id)
    return jsonify(json.loads(char.ataques or '[]'))

@app.route('/api/add_attack/<int:character_id>', methods=['POST'])
@login_required
def add_attack(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    attacks = json.loads(char.ataques or '[]')
    
    new_attack = {
        'id': len(attacks) + 1,
        'nome': data.get('nome', 'Golpe Sem Nome'),
        'pericia': data.get('pericia', 'Luta'), # e.g. Luta, Pontaria, Feitiçaria, or Atributo Puro
        'dano_dados': data.get('dano_dados', '1d6'),
        'dano_attr': data.get('dano_attr', 'forca'), # forca, destreza, none
        'bonus_acerto': int(data.get('bonus_acerto', 0)),
        'bonus_dano': int(data.get('bonus_dano', 0)),
        'critico': data.get('critico', '20 / x2'),
        'alcance': data.get('alcance', 'Corpo a Corpo'),
        'tipo': data.get('tipo', 'Impacto')
    }
    
    attacks.append(new_attack)
    char.ataques = json.dumps(attacks)
    db.session.commit()
    
    return jsonify(attacks)

@app.route('/api/delete_attack/<int:character_id>/<int:attack_id>', methods=['DELETE'])
@login_required
def delete_attack(character_id, attack_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    attacks = json.loads(char.ataques or '[]')
    attacks = [a for a in attacks if a['id'] != attack_id]
    
    char.ataques = json.dumps(attacks)
    db.session.commit()
    
    return jsonify(attacks)

@app.route('/api/get_spells/<int:character_id>')
@login_required
def get_spells(character_id):
    char = Character.query.get_or_404(character_id)
    return jsonify(json.loads(char.feiticos or '[]'))

@app.route('/api/add_spell/<int:character_id>', methods=['POST'])
@login_required
def add_spell(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    spells = json.loads(char.feiticos or '[]')
    
    new_spell = {
        'id': len(spells) + 1,
        'nivel': int(data.get('nivel', 1)),
        'nome': data.get('nome', 'Feitiço Sem Nome'),
        'custo': int(data.get('custo', 2)),
        'acao': data.get('acao', 'Padrão'),
        'alcance': data.get('alcance', 'Pessoal'),
        'duracao': data.get('duracao', 'Instantânea'),
        'dano': data.get('dano', '3d8'),
        'descricao': data.get('descricao', ''),
        'tipo': data.get('tipo', 'Ativo')
    }
    
    spells.append(new_spell)
    char.feiticos = json.dumps(spells)
    db.session.commit()
    
    return jsonify(spells)

@app.route('/api/delete_spell/<int:character_id>/<int:spell_id>', methods=['DELETE'])
@login_required
def delete_spell(character_id, spell_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    spells = json.loads(char.feiticos or '[]')
    spells = [s for s in spells if s['id'] != spell_id]
    
    char.feiticos = json.dumps(spells)
    db.session.commit()
    
    return jsonify(spells)

@app.route('/api/update_spell/<int:character_id>/<int:spell_id>', methods=['POST'])
@login_required
def update_spell(character_id, spell_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    spells = json.loads(char.feiticos or '[]')
    
    for s in spells:
        if s['id'] == spell_id:
            if 'nome' in data: s['nome'] = data['nome']
            if 'custo' in data: s['custo'] = int(data['custo'])
            if 'nivel' in data: s['nivel'] = int(data['nivel'])
            if 'dano' in data: s['dano'] = data['dano']
            if 'equipado' in data: s['equipado'] = bool(data['equipado'])
            if 'ativo' in data: s['ativo'] = bool(data['ativo'])
            if 'tipo' in data: s['tipo'] = data['tipo']
            break
            
    char.feiticos = json.dumps(spells)
    db.session.commit()
    return jsonify(spells)

@app.route('/api/get_summons/<int:character_id>')
@login_required
def get_summons(character_id):
    char = Character.query.get_or_404(character_id)
    return jsonify(json.loads(char.invocacoes or '[]'))

@app.route('/api/add_summon/<int:character_id>', methods=['POST'])
@login_required
def add_summon(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    summons = json.loads(char.invocacoes or '[]')
    
    new_summon = {
        'id': len(summons) + 1,
        'nome': data.get('nome', 'Novo Shikigami'),
        'hp_atual': int(data.get('hp_max', 10)),
        'hp_max': int(data.get('hp_max', 10)),
        'pe_atual': int(data.get('pe_max', 5)),
        'pe_max': int(data.get('pe_max', 5)),
        'ataque': data.get('ataque', '1d6+2'),
        'defesa': int(data.get('defesa', 12)),
        'desc': data.get('desc', '')
    }
    
    summons.append(new_summon)
    char.invocacoes = json.dumps(summons)
    db.session.commit()
    
    return jsonify(summons)

@app.route('/api/update_summon/<int:character_id>/<int:summon_id>', methods=['POST'])
@login_required
def update_summon(character_id, summon_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    summons = json.loads(char.invocacoes or '[]')
    
    for s in summons:
        if s['id'] == summon_id:
            if 'nome' in data: s['nome'] = data['nome']
            if 'hp_atual' in data: s['hp_atual'] = int(data['hp_atual'])
            if 'hp_max' in data: s['hp_max'] = int(data['hp_max'])
            if 'pe_atual' in data: s['pe_atual'] = int(data['pe_atual'])
            if 'pe_max' in data: s['pe_max'] = int(data['pe_max'])
            if 'ataque' in data: s['ataque'] = data['ataque']
            if 'defesa' in data: s['defesa'] = int(data['defesa'])
            if 'desc' in data: s['desc'] = data['desc']
            break
            
    char.invocacoes = json.dumps(summons)
    db.session.commit()
    return jsonify(summons)

@app.route('/api/delete_summon/<int:character_id>/<int:summon_id>', methods=['DELETE'])
@login_required
def delete_summon(character_id, summon_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    summons = json.loads(char.invocacoes or '[]')
    summons = [s for s in summons if s['id'] != summon_id]
    
    char.invocacoes = json.dumps(summons)
    db.session.commit()
    
    return jsonify(summons)

def scale_current_status_proportionally(char, old_pv_max, old_pe_max):
    """
    Recalculates char.status.pv_atual, pe_atual, and integridade_atual
    proportionally based on their new maximums, compared to the old maximums.
    """
    if not char.status:
        return
        
    old_pv_pct = char.status.pv_atual / old_pv_max if old_pv_max > 0 else 1.0
    old_pe_pct = char.status.pe_atual / old_pe_max if old_pe_max > 0 else 1.0
    
    db.session.flush()
    
    new_pv_max = char.status.pv_max
    new_pe_max = char.status.pe_max
    
    char.status.pv_atual = max(0, min(new_pv_max, int(round(old_pv_pct * new_pv_max))))
    char.status.pe_atual = max(0, min(new_pe_max, int(round(old_pe_pct * new_pe_max))))
    char.status.integridade_atual = max(0, min(new_pv_max, int(round(old_pv_pct * new_pv_max))))

def get_character_json(char):
    return {
        'id': char.id,
        'nome': char.nome,
        'origem': char.origem,
        'especializacao': char.especializacao,
        'grau': char.grau,
        'nivel': char.nivel,
        'xp': char.xp,
        'imagem_url': char.imagem_url,
        'cor_energia': char.cor_energia,
        'pontos_atributos': char.pontos_atributos or 0,
        'peso': char.peso or '72kg',
        'altura': char.altura or '1.82m',
        'afiliacao': char.afiliacao or 'Colégio Técnico de Jujutsu',
        'votos_ativos': char.votos_ativos or 'Revelação da Técnica (+2 CD Feitiços)',
        'pericias': json.loads(char.pericias or '{}'),
        'ataques': json.loads(char.ataques or '[]'),
        'resistencias': json.loads(char.resistencias or '{}'),
        'rds': json.loads(char.rds or '{}'),
        'habilidades_talentos': json.loads(char.habilidades_talentos or '[]'),
        'dados_vida': json.loads(char.dados_vida or '{}'),
        'anotacoes': char.anotacoes or '',
        'caracteristicas': json.loads(char.caracteristicas or '[]'),
        'configuracoes': json.loads(char.configuracoes or '{}'),
        'dominio': json.loads(char.dominio or '{}'),
        'inventario': json.loads(char.inventario or '[]'),
        'feiticos': json.loads(char.feiticos or '[]'),
        'invocacoes': json.loads(char.invocacoes or '[]'),
        'iniciativa': char.iniciativa,
        'atencao_passiva': char.atencao_passiva,
        'cd_especializacao': char.cd_especializacao,
        'bonus_corpo_corpo': char.bonus_corpo_corpo,
        'bonus_a_distancia': char.bonus_a_distancia,
        'bonus_amaldicoado': char.bonus_amaldicoado,
        'training_bonus': char.training_bonus,
        'half_level': char.half_level,
        'status': {
            'pv_atual': char.status.pv_atual if char.status else 0,
            'pv_max': char.status.pv_max if char.status else 0,
            'pe_atual': char.status.pe_atual if char.status else 0,
            'pe_max': char.status.pe_max if char.status else 0,
            'integridade_atual': char.status.integridade_atual if char.status else 0,
            'integridade_max': char.status.integridade_max if char.status else 0,
            'estado_alma': char.status.estado_alma if char.status else 'Desconhecido',
            'falhas_morte': char.status.falhas_morte if char.status else 0,
            'sucessos_morte': char.status.sucessos_morte if char.status else 0,
            'defesa': char.status.defesa if char.status else 0
        },
        'attributes': {
            'forca': char.attributes.forca if char.attributes else 0,
            'destreza': char.attributes.destreza if char.attributes else 0,
            'constituicao': char.attributes.constituicao if char.attributes else 0,
            'inteligencia': char.attributes.inteligencia if char.attributes else 0,
            'sabedoria': char.attributes.sabedoria if char.attributes else 0,
            'presenca': char.attributes.presenca if char.attributes else 0
        }
    }

def roll_dice(expression):
    import random
    import re
    expr = expression.strip().lower().replace(" ", "")
    match = re.match(r'^(\d*)d(\d+)([+-]\d+)?$', expr)
    if not match:
        return None, "Expressão inválida"
    
    num_dice = int(match.group(1)) if match.group(1) else 1
    faces = int(match.group(2))
    modifier = int(match.group(3)) if match.group(3) else 0
    
    rolls = [random.randint(1, faces) for _ in range(num_dice)]
    total = sum(rolls) + modifier
    
    rolls_str = " + ".join(map(str, rolls))
    if modifier > 0:
        rolls_str += f" + {modifier}"
    elif modifier < 0:
        rolls_str += f" - {abs(modifier)}"
        
    return total, f"({rolls_str}) = {total}"

@app.route('/api/update_resistencias/<int:character_id>', methods=['POST'])
@login_required
def update_resistencias(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()
    if not data or 'resistencias' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    char.resistencias = json.dumps(data['resistencias'])
    db.session.commit()
    return jsonify({
        'message': 'Resistances updated successfully',
        'character': get_character_json(char)
    })

@app.route('/api/update_rds/<int:character_id>', methods=['POST'])
@login_required
def update_rds(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()
    if not data or 'rds' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    char.rds = json.dumps(data['rds'])
    db.session.commit()
    return jsonify({
        'message': 'RDs updated successfully',
        'character': get_character_json(char)
    })

@app.route('/api/update_configuracoes/<int:character_id>', methods=['POST'])
@login_required
def update_configuracoes(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()
    if not data or 'configuracoes' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    
    try:
        config = json.loads(char.configuracoes or '{}')
    except:
        config = {}
        
    config.update(data['configuracoes'])
    char.configuracoes = json.dumps(config)
    db.session.commit()
    return jsonify({
        'message': 'Configuracoes updated successfully',
        'character': get_character_json(char)
    })

@app.route('/api/update_dados_vida/<int:character_id>', methods=['POST'])
@login_required
def update_dados_vida(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()
    if not data or 'dados_vida' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    char.dados_vida = json.dumps(data['dados_vida'])
    db.session.commit()
    return jsonify({
        'message': 'Dados de vida updated successfully',
        'character': get_character_json(char)
    })

@app.route('/api/update_anotacoes/<int:character_id>', methods=['POST'])
@login_required
def update_anotacoes(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()
    if not data or 'anotacoes' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    char.anotacoes = data['anotacoes']
    db.session.commit()
    return jsonify({
        'message': 'Anotacoes updated successfully',
        'character': get_character_json(char)
    })

@app.route('/api/add_talent/<int:character_id>', methods=['POST'])
@login_required
def add_talent(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json() or {}
    talents = json.loads(char.habilidades_talentos or '[]')
    
    new_talent = {
        'id': len(talents) + 1,
        'nome': data.get('nome', 'Novo Talento'),
        'tipo': data.get('tipo', 'Classe'),
        'custo': int(data.get('custo', 0)),
        'execucao': data.get('execucao', 'Ação Padrão'),
        'alcance': data.get('alcance', 'Pessoal'),
        'duracao': data.get('duracao', 'Instantânea'),
        'descricao': data.get('descricao', ''),
        'dado_rolagem': data.get('dado_rolagem', '')
    }
    
    talents.append(new_talent)
    char.habilidades_talentos = json.dumps(talents)
    db.session.commit()
    return jsonify(talents)

@app.route('/api/update_talent/<int:character_id>/<int:talent_id>', methods=['POST'])
@login_required
def update_talent(character_id, talent_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json() or {}
    talents = json.loads(char.habilidades_talentos or '[]')
    
    for t in talents:
        if t['id'] == talent_id:
            if 'nome' in data: t['nome'] = data['nome']
            if 'tipo' in data: t['tipo'] = data['tipo']
            if 'custo' in data: t['custo'] = int(data['custo'])
            if 'execucao' in data: t['execucao'] = data['execucao']
            if 'alcance' in data: t['alcance'] = data['alcance']
            if 'duracao' in data: t['duracao'] = data['duracao']
            if 'descricao' in data: t['descricao'] = data['descricao']
            if 'dado_rolagem' in data: t['dado_rolagem'] = data['dado_rolagem']
            break
            
    char.habilidades_talentos = json.dumps(talents)
    db.session.commit()
    return jsonify(talents)

@app.route('/api/delete_talent/<int:character_id>/<int:talent_id>', methods=['DELETE'])
@login_required
def delete_talent(character_id, talent_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    talents = json.loads(char.habilidades_talentos or '[]')
    talents = [t for t in talents if t['id'] != talent_id]
    
    char.habilidades_talentos = json.dumps(talents)
    db.session.commit()
    return jsonify(talents)

@app.route('/api/use_talent/<int:character_id>/<int:talent_id>', methods=['POST'])
@login_required
def use_talent(character_id, talent_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    if not char.status:
        return jsonify({'error': 'Character status not initialized'}), 400
        
    talents = json.loads(char.habilidades_talentos or '[]')
    talent = next((t for t in talents if t['id'] == talent_id), None)
    if not talent:
        return jsonify({'error': 'Talent not found'}), 404
        
    cost = int(talent.get('custo', 0))
    if char.status.pe_atual < cost:
        return jsonify({'error': 'Pontos de Energia (PE) insuficientes!'}), 400
        
    char.status.pe_atual -= cost
    db.session.commit()
    
    roll_result = None
    roll_desc = ""
    expr = talent.get('dado_rolagem', '')
    if expr:
        try:
            val, desc = roll_dice(expr)
            if val is not None:
                roll_result = val
                roll_desc = desc
        except Exception as e:
            roll_desc = f"Erro ao rolar: {str(e)}"
            
    return jsonify({
        'message': f"Usou o talento {talent['nome']}",
        'talent_nome': talent['nome'],
        'pe_atual': char.status.pe_atual,
        'cost': cost,
        'roll_result': roll_result,
        'roll_desc': roll_desc,
        'descricao': talent.get('descricao', '')
    })

@app.route('/api/use_attack/<int:character_id>/<int:attack_id>', methods=['POST'])
@login_required
def use_attack(character_id, attack_id):
    import random
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    if not char.status:
        return jsonify({'error': 'Character status not initialized'}), 400
        
    attacks = json.loads(char.ataques or '[]')
    attack = next((a for a in attacks if a['id'] == attack_id), None)
    if not attack:
        return jsonify({'error': 'Ataque não encontrado'}), 404
        
    data = request.get_json() or {}
    target_id = data.get('target_id')
    target_char = None
    if target_id:
        target_char = Character.query.get(int(target_id))
        if not target_char or not target_char.status:
            return jsonify({'error': 'Alvo inválido ou sem status'}), 400
            
    # Roll Attack Check (Acerto)
    d20_roll = random.randint(1, 20)
    bonus_acerto = int(attack.get('bonus_acerto', 0))
    total_acerto = d20_roll + bonus_acerto
    
    is_crit = False
    crit_range = 20
    crit_str = attack.get('critico', '20 / x2')
    # Try to parse critical range from something like "19 / x2" or "19"
    try:
        parts = crit_str.split('/')
        val = int(parts[0].strip())
        if 1 <= val <= 20:
            crit_range = val
    except:
        pass
        
    if d20_roll >= crit_range:
        is_crit = True
        
    hit = True
    miss_reason = ""
    target_defesa = 10
    
    if target_char:
        target_defesa = target_char.status.defesa
        if d20_roll == 20:
            hit = True
        elif d20_roll == 1:
            hit = False
            miss_reason = "Falha Crítica (1 natural)"
        elif total_acerto >= target_defesa:
            hit = True
        else:
            hit = False
            miss_reason = f"Rolagem ({total_acerto}) menor que a Defesa do alvo ({target_defesa})"
            
    damage_roll_total = 0
    damage_roll_desc = ""
    bonus_dano = int(attack.get('bonus_dano', 0))
    attr_mod = 0
    dano_attr = attack.get('dano_attr', '')
    if dano_attr == 'forca':
        attr_mod = char.mod_forca
    elif dano_attr == 'destreza':
        attr_mod = char.mod_destreza
        
    total_bonus = bonus_dano + attr_mod
    
    final_damage = 0
    rd_applied = 0
    damage_type = attack.get('tipo', 'Impacto')
    
    rd_map = {
        'corte': 'COR', 'imp': 'IMP', 'impacto': 'IMP', 'perf': 'PER', 'perfuracao': 'PER', 'perfuração': 'PER',
        'con': 'CON', 'congelamento': 'CON', 'que': 'QUE', 'queimadura': 'QUE', 'fogo': 'QUE',
        'cho': 'CHO', 'choque': 'CHO', 'eletricidade': 'CHO', 'son': 'SON', 'sonico': 'SON', 'sônico': 'SON',
        'ene': 'ENE', 'energia': 'ENE', 'amaldiçoada': 'ENE', 'amaldicoada': 'ENE', 'psi': 'PSI', 'mental': 'PSI',
        'rad': 'RAD', 'radiacao': 'RAD', 'radiação': 'RAD', 'nec': 'NEC', 'necrotico': 'NEC', 'necrótico': 'NEC',
        'ven': 'VEN', 'veneno': 'VEN', 'acido': 'VEN', 'ácido': 'VEN'
    }
    
    if hit:
        dano_dados = attack.get('dano_dados', '1d6')
        try:
            val, desc = roll_dice(dano_dados)
            if val is not None:
                damage_roll_total = val
                damage_roll_desc = desc
            else:
                damage_roll_total = random.randint(1, 6)
                damage_roll_desc = f"(1d6 no fallback) = {damage_roll_total}"
        except Exception as e:
            damage_roll_total = random.randint(1, 6)
            damage_roll_desc = f"(1d6 no fallback devido a erro {str(e)}) = {damage_roll_total}"
            
        base_dmg = damage_roll_total
        if is_crit:
            # Double only the rolled dice as standard JJK rules or double whole roll
            base_dmg = damage_roll_total * 2
            damage_roll_desc = f"CRÍTICO! [ {damage_roll_desc} ] x2 = {base_dmg}"
            
        total_damage = base_dmg + total_bonus
        
        if target_char:
            target_rds = json.loads(target_char.rds or '{}')
            rd_abbrev = rd_map.get(damage_type.lower(), 'IMP')
            rd_applied = int(target_rds.get(rd_abbrev, 0))
            final_damage = max(0, total_damage - rd_applied)
            
            # Deduct target PV
            old_pv = target_char.status.pv_atual
            target_char.status.pv_atual = max(0, target_char.status.pv_atual - final_damage)
            pv_lost = old_pv - target_char.status.pv_atual
        else:
            final_damage = total_damage
            
    # Log generation
    time_str = time.strftime('%H:%M')
    log_title = "Ataque Realizado"
    if target_char:
        if hit:
            crit_badge = " [CRÍTICO]" if is_crit else ""
            log_content = (
                f"<b>{char.nome}</b> atacou <b>{target_char.nome}</b> com <b>{attack['nome']}</b> e <b>ACERTOU</b>!{crit_badge}<br>"
                f"[Ataque] Acerto: d20 ({d20_roll}) + {bonus_acerto} = {total_acerto} vs Defesa {target_defesa}<br>"
                f"[Dano] Dano: {damage_roll_desc} + Bônus ({total_bonus}) = {total_damage} ({damage_type})<br>"
                f"[RD] RD Alvo ({rd_abbrev}): -{rd_applied}<br>"
                f"[Dano Final] Dano Final sofrido por {target_char.nome}: <b>{final_damage} PV</b>"
            )
        else:
            log_content = (
                f"<b>{char.nome}</b> atacou <b>{target_char.nome}</b> com <b>{attack['nome']}</b> mas <b>ERROU</b>!<br>"
                f"[Ataque] Acerto: d20 ({d20_roll}) + {bonus_acerto} = {total_acerto} vs Defesa {target_defesa}<br>"
                f"[Falha] Motivo: {miss_reason}"
            )
    else:
        # Rolou sem alvo específico
        crit_badge = " [CRÍTICO]" if is_crit else ""
        if hit:
            log_content = (
                f"<b>{char.nome}</b> atacou com <b>{attack['nome']}</b>!{crit_badge}<br>"
                f"[Ataque] Acerto: d20 ({d20_roll}) + {bonus_acerto} = {total_acerto}<br>"
                f"[Dano] Dano: {damage_roll_desc} + Bônus ({total_bonus}) = {total_damage} ({damage_type})"
            )
            
    # Append log to caster
    try:
        caster_logs = json.loads(char.recent_logs or '[]')
        caster_logs.insert(0, {'title': log_title, 'content': log_content, 'time': time_str, 'timestamp': time.time()})
        char.recent_logs = json.dumps(caster_logs[:15])
    except:
        pass
        
    # Append log to target if exists
    if target_char:
        try:
            target_logs = json.loads(target_char.recent_logs or '[]')
            target_logs.insert(0, {'title': log_title, 'content': log_content, 'time': time_str, 'timestamp': time.time()})
            target_char.recent_logs = json.dumps(target_logs[:15])
        except:
            pass
            
    db.session.commit()
    
    return jsonify({
        'hit': hit,
        'is_crit': is_crit,
        'd20_roll': d20_roll,
        'total_acerto': total_acerto,
        'damage_roll_desc': damage_roll_desc,
        'total_damage': damage_roll_total + total_bonus,
        'rd_applied': rd_applied,
        'final_damage': final_damage,
        'caster_pe': char.status.pe_atual,
        'target_pv': target_char.status.pv_atual if target_char else None,
        'log_content': log_content
    })

@app.route('/api/use_spell/<int:character_id>/<int:spell_id>', methods=['POST'])
@login_required
def use_spell(character_id, spell_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    if not char.status:
        return jsonify({'error': 'Character status not initialized'}), 400
        
    spells = json.loads(char.feiticos or '[]')
    spell = next((s for s in spells if s['id'] == spell_id), None)
    if not spell:
        return jsonify({'error': 'Feitiço não encontrado'}), 404
        
    cost = int(spell.get('custo', 2))
    if char.status.pe_atual < cost:
        return jsonify({'error': f'Energia Amaldiçoada (PE) insuficiente! Custo: {cost} PE, Atual: {char.status.pe_atual} PE'}), 400
        
    data = request.get_json() or {}
    target_id = data.get('target_id')
    target_char = None
    if target_id:
        target_char = Character.query.get(int(target_id))
        if not target_char or not target_char.status:
            return jsonify({'error': 'Alvo inválido ou sem status'}), 400
            
    char.status.pe_atual -= cost
    
    # Roll spell damage or healing
    dano_expr = spell.get('dano', '')
    damage_roll_total = 0
    damage_roll_desc = ""
    
    is_healing = False
    name_lower = spell['nome'].lower()
    desc_lower = spell.get('descricao', '').lower()
    if 'cura' in name_lower or 'curar' in name_lower or 'recupera' in name_lower or 'cura' in desc_lower:
        is_healing = True
        
    if dano_expr:
        try:
            val, desc = roll_dice(dano_expr)
            if val is not None:
                damage_roll_total = val
                damage_roll_desc = desc
            else:
                damage_roll_total = 0
                damage_roll_desc = "N/A"
        except Exception as e:
            damage_roll_desc = f"Erro: {str(e)}"
            
    final_effect = 0
    rd_applied = 0
    rd_abbrev = "ENE"
    
    rd_map = {
        'corte': 'COR', 'imp': 'IMP', 'impacto': 'IMP', 'perf': 'PER', 'perfuracao': 'PER', 'perfuração': 'PER',
        'con': 'CON', 'congelamento': 'CON', 'que': 'QUE', 'queimadura': 'QUE', 'fogo': 'QUE',
        'cho': 'CHO', 'choque': 'CHO', 'eletricidade': 'CHO', 'son': 'SON', 'sonico': 'SON', 'sônico': 'SON',
        'ene': 'ENE', 'energia': 'ENE', 'amaldiçoada': 'ENE', 'amaldicoada': 'ENE', 'psi': 'PSI', 'mental': 'PSI',
        'rad': 'RAD', 'radiacao': 'RAD', 'radiação': 'RAD', 'nec': 'NEC', 'necrotico': 'NEC', 'necrótico': 'NEC',
        'ven': 'VEN', 'veneno': 'VEN', 'acido': 'VEN', 'ácido': 'VEN'
    }
    
    time_str = time.strftime('%H:%M')
    log_title = "Técnica Conjurada"
    
    if target_char:
        if is_healing:
            old_pv = target_char.status.pv_atual
            target_char.status.pv_atual = min(target_char.status.pv_max, target_char.status.pv_atual + damage_roll_total)
            healed = target_char.status.pv_atual - old_pv
            final_effect = healed
            log_content = (
                f"[Ritual] <b>{char.nome}</b> conjurou <b>{spell['nome']}</b> em <b>{target_char.nome}</b>!<br>"
                f"Custo: {cost} PE (Caster agora tem {char.status.pe_atual} PE)<br>"
                f"Cura: {damage_roll_desc}<br>"
                f"<b>{target_char.nome}</b> recuperou <b>{healed} PV</b>!"
            )
        else:
            # Deal damage, default to Cursed Energy (ENE) or custom
            target_rds = json.loads(target_char.rds or '{}')
            # Look up if description or name matches any specific damage type
            matched_type = 'ENE'
            for t in rd_map.keys():
                if t in name_lower or t in desc_lower:
                    matched_type = rd_map[t]
                    break
            rd_abbrev = matched_type
            rd_applied = int(target_rds.get(rd_abbrev, 0))
            final_damage = max(0, damage_roll_total - rd_applied)
            final_effect = final_damage
            
            old_pv = target_char.status.pv_atual
            target_char.status.pv_atual = max(0, target_char.status.pv_atual - final_damage)
            
            log_content = (
                f"[Ritual] <b>{char.nome}</b> conjurou <b>{spell['nome']}</b> em <b>{target_char.nome}</b>!<br>"
                f"Custo: {cost} PE (Caster agora tem {char.status.pe_atual} PE)<br>"
                f"Dano Rolado: {damage_roll_desc} ({rd_abbrev})<br>"
                f"[RD] RD Alvo ({rd_abbrev}): -{rd_applied}<br>"
                f"[Dano Final] Dano Final sofrido por {target_char.nome}: <b>{final_damage} PV</b>"
            )
    else:
        # Casted on self / no target
        log_content = (
            f"[Ritual] <b>{char.nome}</b> conjurou <b>{spell['nome']}</b>!<br>"
            f"Custo: {cost} PE (Caster agora tem {char.status.pe_atual} PE)<br>"
            f"Descrição: <i>{spell.get('descricao', 'Sem descrição')}</i>"
        )
        if dano_expr:
            log_content += f"<br>Efeito Rolado: {damage_roll_desc}"
            
    # Append log to caster
    try:
        caster_logs = json.loads(char.recent_logs or '[]')
        caster_logs.insert(0, {'title': log_title, 'content': log_content, 'time': time_str, 'timestamp': time.time()})
        char.recent_logs = json.dumps(caster_logs[:15])
    except:
        pass
        
    # Append log to target if exists
    if target_char:
        try:
            target_logs = json.loads(target_char.recent_logs or '[]')
            target_logs.insert(0, {'title': log_title, 'content': log_content, 'time': time_str, 'timestamp': time.time()})
            target_char.recent_logs = json.dumps(target_logs[:15])
        except:
            pass
            
    db.session.commit()
    
    return jsonify({
        'pe_atual': char.status.pe_atual,
        'cost': cost,
        'damage_roll_total': damage_roll_total,
        'damage_roll_desc': damage_roll_desc,
        'final_effect': final_effect,
        'rd_applied': rd_applied,
        'target_pv': target_char.status.pv_atual if target_char else None,
        'is_healing': is_healing,
        'log_content': log_content
    })

@app.route('/api/use_lobby_talent/<int:character_id>/<int:talent_id>', methods=['POST'])
@login_required
def use_lobby_talent(character_id, talent_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    if not char.status:
        return jsonify({'error': 'Character status not initialized'}), 400
        
    talents = json.loads(char.habilidades_talentos or '[]')
    talent = next((t for t in talents if t['id'] == talent_id), None)
    if not talent:
        return jsonify({'error': 'Talento não encontrado'}), 404
        
    cost = int(talent.get('custo', 0))
    if char.status.pe_atual < cost:
        return jsonify({'error': f'Pontos de Energia (PE) insuficientes! Custo: {cost} PE, Atual: {char.status.pe_atual} PE'}), 400
        
    data = request.get_json() or {}
    target_id = data.get('target_id')
    target_char = None
    if target_id:
        target_char = Character.query.get(int(target_id))
        if not target_char or not target_char.status:
            return jsonify({'error': 'Alvo inválido ou sem status'}), 400
            
    char.status.pe_atual -= cost
    
    roll_result = None
    roll_desc = ""
    expr = talent.get('dado_rolagem', '')
    if expr:
        try:
            val, desc = roll_dice(expr)
            if val is not None:
                roll_result = val
                roll_desc = desc
        except Exception as e:
            roll_desc = f"Erro: {str(e)}"
            
    time_str = time.strftime('%H:%M')
    log_title = "Habilidade Usada"
    
    if target_char:
        log_content = (
            f"[Talento] <b>{char.nome}</b> ativou <b>{talent['nome']}</b> em <b>{target_char.nome}</b>!<br>"
            f"Custo: {cost} PE (Usuário agora tem {char.status.pe_atual} PE)<br>"
            f"Descrição: <i>{talent.get('descricao', 'Sem descrição')}</i>"
        )
        if expr:
            log_content += f"<br>Efeito Rolado: {roll_desc}"
    else:
        log_content = (
            f"[Talento] <b>{char.nome}</b> ativou <b>{talent['nome']}</b>!<br>"
            f"Custo: {cost} PE (Usuário agora tem {char.status.pe_atual} PE)<br>"
            f"Descrição: <i>{talent.get('descricao', 'Sem descrição')}</i>"
        )
        if expr:
            log_content += f"<br>Efeito Rolado: {roll_desc}"
            
    # Append log to caster
    try:
        caster_logs = json.loads(char.recent_logs or '[]')
        caster_logs.insert(0, {'title': log_title, 'content': log_content, 'time': time_str, 'timestamp': time.time()})
        char.recent_logs = json.dumps(caster_logs[:15])
    except:
        pass
        
    # Append log to target if exists
    if target_char:
        try:
            target_logs = json.loads(target_char.recent_logs or '[]')
            target_logs.insert(0, {'title': log_title, 'content': log_content, 'time': time_str, 'timestamp': time.time()})
            target_char.recent_logs = json.dumps(target_logs[:15])
        except:
            pass
            
    db.session.commit()
    
    return jsonify({
        'pe_atual': char.status.pe_atual,
        'cost': cost,
        'roll_result': roll_result,
        'roll_desc': roll_desc,
        'target_pv': target_char.status.pv_atual if target_char else None,
        'log_content': log_content
    })

@app.route('/api/rest_character/<int:character_id>', methods=['POST'])
@login_required
def rest_character(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    if not char.status:
        return jsonify({'error': 'Character status not initialized'}), 400
        
    data = request.get_json() or {}
    tipo = data.get('tipo', 'longo')
    
    try:
        dv = json.loads(char.dados_vida or '{}')
    except:
        dv = {}
        
    for d in ['d8', 'd10', 'd12']:
        if d not in dv:
            dv[d] = {'gastos': 0, 'max': 0}
            
    log_message = ""
    roll_desc = ""
    pv_recuperado = 0
    pe_recuperado = 0
    
    if tipo == 'curto':
        dice_type = data.get('dice_type')
        if not dice_type or dice_type not in dv:
            return jsonify({'error': 'Tipo de dado inválido'}), 400
            
        d_info = dv[dice_type]
        if d_info['max'] <= 0:
            return jsonify({'error': f'Você não possui dados de vida do tipo {dice_type}'}), 400
        if d_info['gastos'] >= d_info['max']:
            return jsonify({'error': f'Todos os dados de vida {dice_type} já foram gastos'}), 400
            
        d_info['gastos'] += 1
        char.dados_vida = json.dumps(dv)
        
        import random
        faces = int(dice_type.replace('d', ''))
        die_roll = random.randint(1, faces)
        con_mod = char.mod_constituicao
        total_heal = max(1, die_roll + con_mod)
        
        old_pv = char.status.pv_atual
        char.status.pv_atual = min(char.status.pv_max, char.status.pv_atual + total_heal)
        pv_recuperado = char.status.pv_atual - old_pv
        
        roll_desc = f"1{dice_type} ({die_roll}) + CON ({con_mod}) = {total_heal} PV"
        log_message = f"Realizou um Descanso Curto gastando 1{dice_type} e recuperando {pv_recuperado} PV!"
        
    elif tipo == 'longo':
        old_pv = char.status.pv_atual
        char.status.pv_atual = char.status.pv_max
        pv_recuperado = char.status.pv_max - old_pv
        
        old_pe = char.status.pe_atual
        char.status.pe_atual = char.status.pe_max
        pe_recuperado = char.status.pe_max - old_pe
        
        char.status.integridade_atual = char.status.integridade_max
        
        for d, d_info in dv.items():
            if d_info['max'] > 0:
                recover_amount = max(1, d_info['max'] // 2)
                d_info['gastos'] = max(0, d_info['gastos'] - recover_amount)
                
        char.dados_vida = json.dumps(dv)
        log_message = f"Realizou um Descanso Longo! Recuperou {pv_recuperado} PV, {pe_recuperado} PE, Integridade da Alma e metade dos Dados de Vida!"
        
    db.session.commit()
    return jsonify({
        'message': log_message,
        'pv_atual': char.status.pv_atual,
        'pv_max': char.status.pv_max,
        'pe_atual': char.status.pe_atual,
        'pe_max': char.status.pe_max,
        'integridade_atual': char.status.integridade_atual,
        'dados_vida': dv,
        'roll_desc': roll_desc,
        'pv_recuperado': pv_recuperado,
        'pe_recuperado': pe_recuperado,
        'character': get_character_json(char)
    })

@app.route('/api/upload_avatar/<int:character_id>', methods=['POST'])
@login_required
def upload_avatar(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    if file:
        filename = secure_filename(file.filename)
        ext = os.path.splitext(filename)[1]
        name = os.path.splitext(filename)[0]
        unique_filename = f"{name}_{int(time.time())}{ext}"
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        char.imagem_url = f"/static/uploads/{unique_filename}"
        db.session.commit()
        
        return jsonify({
            'message': 'Avatar uploaded successfully',
            'imagem_url': char.imagem_url,
            'character': get_character_json(char)
        })
        
    return jsonify({'error': 'Upload failed'}), 400

def _process_excel_import(char, wb, filename):
    def safe_int(val, default=0):
        if val is None: return default
        try: return int(float(val))
        except: return default

    def safe_float(val, default=0.0):
        if val is None: return default
        try: return float(val)
        except: return default

    def safe_str(val, default=""):
        if val is None: return default
        return str(val).strip()

    sheet_f = wb['Ficha Pessoal']
    
    # Correct cell coordinates from actual JJK 2.5 Excel Model
    nome_val = safe_str(sheet_f['S2'].value)
    if not nome_val or nome_val.lower() == "nome":
        nome_val = char.nome
    char.nome = nome_val
    
    origem_val = safe_str(sheet_f['AH2'].value)
    if origem_val and origem_val.lower() != "origem":
        char.origem = origem_val
        
    espec_val = safe_str(sheet_f['AH3'].value)
    if espec_val and espec_val.lower() not in ("especialização", "especializacao"):
        char.especializacao = espec_val
        
    grau_val = safe_str(sheet_f['AU3'].value)
    if grau_val and grau_val.lower() != "grau":
        char.grau = grau_val
        
    nivel_val = sheet_f['BD3'].value
    if nivel_val is not None:
        char.nivel = max(1, safe_int(nivel_val, char.nivel))
        
    xp_val = sheet_f['AU4'].value
    if xp_val is not None:
        char.xp = safe_int(xp_val, char.xp)
        
    # Read extra health/defense config checkboxes from spreadsheet
    try:
        config = json.loads(char.configuracoes or '{}')
    except:
        config = {}
    config['pv_kamo'] = bool(sheet_f['AA7'].value)
    config['pv_robustez'] = bool(sheet_f['AA8'].value)
    config['pv_deslocamento'] = bool(sheet_f['AA9'].value)
    config['pv_vigor_infinito'] = bool(sheet_f['AA10'].value)
    config['pv_outros'] = safe_int(sheet_f['Z11'].value, 0)
    
    def map_attr_str(val):
        if not val: return 'forca'
        v = str(val).strip().lower()
        if 'for' in v: return 'forca'
        if 'des' in v: return 'destreza'
        if 'con' in v: return 'constituicao'
        if 'int' in v: return 'inteligencia'
        if 'sab' in v: return 'sabedoria'
        if 'pre' in v: return 'presenca'
        return 'forca'

    config['ataque_corpo_corpo'] = {
        'treinada': bool(sheet_f['O18'].value),
        'outros': safe_int(sheet_f['P18'].value, 0),
        'atributo': map_attr_str(sheet_f['O20'].value)
    }
    config['ataque_a_distancia'] = {
        'treinada': bool(sheet_f['R18'].value),
        'outros': safe_int(sheet_f['S18'].value, 0),
        'atributo': map_attr_str(sheet_f['R20'].value)
    }
    config['ataque_amaldicoado'] = {
        'treinada': bool(sheet_f['U18'].value),
        'outros': safe_int(sheet_f['V18'].value, 0),
        'atributo': map_attr_str(sheet_f['U20'].value)
    }
    
    char.configuracoes = json.dumps(config)
    
    if char.attributes:
        char.attributes.forca = safe_int(sheet_f['B10'].value, 10)
        char.attributes.destreza = safe_int(sheet_f['F10'].value, 10)
        char.attributes.constituicao = safe_int(sheet_f['J10'].value, 10)
        char.attributes.inteligencia = safe_int(sheet_f['B14'].value, 10)
        char.attributes.sabedoria = safe_int(sheet_f['F14'].value, 10)
        char.attributes.presenca = safe_int(sheet_f['J14'].value, 10)
    
    pericias = json.loads(char.pericias or '{}')
    
    for r in range(25, 36):
        s_name = safe_str(sheet_f.cell(row=r, column=2).value)
        if s_name:
            clean_name = s_name.split(' (')[0].strip()
            matched_key = None
            for k in pericias.keys():
                if k.lower() == clean_name.lower() or k.lower() == s_name.lower():
                    matched_key = k
                    break
            
            if matched_key:
                is_trained = bool(sheet_f.cell(row=r, column=10).value)
                is_master = bool(sheet_f.cell(row=r, column=11).value)
                pericias[matched_key]['treinada'] = is_trained
                pericias[matched_key]['mestre'] = is_master

    for r in range(25, 36):
        s_name = safe_str(sheet_f.cell(row=r, column=14).value)
        if s_name:
            clean_name = s_name.split(' (')[0].strip()
            matched_key = None
            for k in pericias.keys():
                if k.lower() == clean_name.lower() or k.lower() == s_name.lower():
                    matched_key = k
                    break
            
            if matched_key:
                is_trained = bool(sheet_f.cell(row=r, column=22).value)
                is_master = bool(sheet_f.cell(row=r, column=23).value) if sheet_f.cell(row=r, column=23).value is not None else False
                pericias[matched_key]['treinada'] = is_trained
                pericias[matched_key]['mestre'] = is_master

    resistencias = json.loads(char.resistencias or '{}')
    res_mappings = {
        'Astúcia': ('Y18', 'Z18'),
        'Fortitude': ('AB18', 'AC18'),
        'Integridade': ('AE18', 'AF18'),
        'Reflexos': ('AH18', 'AI18'),
        'Vontade': ('AK18', 'AL18')
    }
    for name, (t_coord, m_coord) in res_mappings.items():
        if name in resistencias:
            resistencias[name]['treinada'] = bool(sheet_f[t_coord].value)
            resistencias[name]['mestre'] = bool(sheet_f[m_coord].value)
    char.resistencias = json.dumps(resistencias)

    rds = json.loads(char.rds or '{}')
    rd_types = ['COR', 'IMP', 'PER', 'CON', 'QUE', 'CHO', 'SON', 'ENE', 'PSI', 'RAD', 'NEC', 'VEN']
    for i, rd_type in enumerate(rd_types):
        col_idx = 27 + i
        rd_val = safe_int(sheet_f.cell(row=22, column=col_idx).value, 0)
        rds[rd_type] = rd_val
    char.rds = json.dumps(rds)

    attacks = []
    for r in range(27, 32):
        w_name = safe_str(sheet_f.cell(row=r, column=27).value)
        if w_name:
            atk_dict = {
                'id': len(attacks) + 1,
                'nome': w_name,
                'pericia': 'Luta' if 'corpo' in safe_str(sheet_f.cell(row=r, column=43).value).lower() else 'Pontaria',
                'dano_dados': safe_str(sheet_f.cell(row=r, column=34).value, '1d6'),
                'dano_attr': 'forca' if 'corpo' in safe_str(sheet_f.cell(row=r, column=43).value).lower() else 'destreza',
                'bonus_acerto': safe_int(sheet_f.cell(row=r, column=32).value, 0),
                'bonus_dano': 0,
                'critico': safe_str(sheet_f.cell(row=r, column=38).value, '20 / x2'),
                'alcance': safe_str(sheet_f.cell(row=r, column=43).value, 'Corpo a Corpo'),
                'tipo': safe_str(sheet_f.cell(row=r, column=40).value, 'Impacto')
            }
            attacks.append(atk_dict)
    char.ataques = json.dumps(attacks)

    talents = []
    for r in range(9, 36):
        t_name = safe_str(sheet_f.cell(row=r, column=60).value)
        if t_name and t_name not in ('Nome', 'NOME'):
            t_atual = safe_str(sheet_f.cell(row=r, column=66).value)
            t_max = safe_str(sheet_f.cell(row=r, column=68).value)
            t_cost = safe_int(sheet_f.cell(row=r, column=70).value, 0)
            
            desc_extra = ""
            if t_atual or t_max:
                desc_extra = f" Usos: {t_atual or '0'}/{t_max or '0'}."
            
            talent_dict = {
                'id': len(talents) + 1,
                'nome': t_name,
                'tipo': 'Classe',
                'custo': t_cost,
                'execucao': 'Ação Padrão',
                'alcance': 'Pessoal',
                'duracao': 'Instantânea',
                'descricao': f"Habilidade importada da planilha Excel.{desc_extra}",
                'dado_rolagem': ''
            }
            talents.append(talent_dict)
    char.habilidades_talentos = json.dumps(talents)

    sheet_i = wb['Registro e Inventário']
    inventory = []
    for r in range(8, 34):
        item_name_1 = safe_str(sheet_i.cell(row=r, column=25).value)
        if item_name_1:
            item_qty_1 = safe_int(sheet_i.cell(row=r, column=33).value, 1)
            item_weight_1 = safe_float(sheet_i.cell(row=r, column=35).value, 0.0)
            inventory.append({
                'id': len(inventory) + 1,
                'nome': item_name_1,
                'qtd': item_qty_1,
                'peso': item_weight_1
            })
        item_name_2 = safe_str(sheet_i.cell(row=r, column=40).value)
        if item_name_2:
            item_qty_2 = safe_int(sheet_i.cell(row=r, column=48).value, 1)
            item_weight_2 = safe_float(sheet_i.cell(row=r, column=50).value, 0.0)
            inventory.append({
                'id': len(inventory) + 1,
                'nome': item_name_2,
                'qtd': item_qty_2,
                'peso': item_weight_2
            })
    char.inventario = json.dumps(inventory)

    sheet_p = wb['Perfil Amaldiçoado']
    spells = []
    
    def parse_spell_str(text, lvl):
        import re
        text_clean = text.strip()
        cost = 2
        match = re.search(r'\((\d+)\s*PE\)', text_clean, re.IGNORECASE)
        if match:
            cost = int(match.group(1))
            text_clean = re.sub(r'\((\d+)\s*PE\)', '', text_clean).strip()
        
        return {
            'id': len(spells) + 1,
            'nivel': lvl,
            'nome': text_clean,
            'custo': cost,
            'acao': 'Padrão',
            'alcance': 'Pessoal',
            'duracao': 'Instantânea',
            'dano': '1d6',
            'descricao': 'Feitiço importado do Excel.'
        }

    spell_blocks = [
        (0, 27, 8, 17),
        (1, 36, 8, 17),
        (2, 45, 8, 17),
        (3, 27, 19, 28),
        (4, 36, 19, 28),
        (5, 45, 19, 28)
    ]
    
    for lvl, col_idx, r_start, r_end in spell_blocks:
        for r in range(r_start, r_end + 1):
            spell_val = safe_str(sheet_p.cell(row=r, column=col_idx).value)
            if spell_val:
                spells.append(parse_spell_str(spell_val, lvl))
    char.feiticos = json.dumps(spells)

    dom_nome = safe_str(sheet_p['BB7'].value)
    if dom_nome:
        char.dominio = json.dumps({
            'nome': dom_nome,
            'tipo': safe_str(sheet_p['BB10'].value, 'Letal'),
            'custo': 20,
            'descricao': safe_str(sheet_p['BB12'].value, 'Feito sob medida em barreira.')
        })

    summons = []
    if len(wb.sheetnames) > 4:
        sheet_s = wb[wb.sheetnames[4]]
        summon_cols = [2, 17, 32]
        for col in summon_cols:
            sum_name = safe_str(sheet_s.cell(row=5, column=col).value)
            if sum_name and "Invoca" not in sum_name:
                sum_hp_max = safe_int(sheet_s.cell(row=10, column=col + 4).value, 10)
                sum_hp_act = safe_int(sheet_s.cell(row=10, column=col + 1).value, sum_hp_max)
                sum_def = safe_int(sheet_s.cell(row=10, column=col + 7).value, 10)
                
                summons.append({
                    'id': len(summons) + 1,
                    'nome': sum_name,
                    'hp_atual': sum_hp_act,
                    'hp_max': sum_hp_max,
                    'pe_atual': 5,
                    'pe_max': 5,
                    'ataque': '1d6+2',
                    'defesa': sum_def,
                    'desc': f"Grau: {safe_str(sheet_s.cell(row=7, column=col + 1).value)}\nDeslocamento: {safe_str(sheet_s.cell(row=10, column=col + 10).value)}"
                })
    char.invocacoes = json.dumps(summons)

    if len(wb.sheetnames) > 5:
        sheet_t = wb[wb.sheetnames[5]]
        reforco_count = 0
        for r in range(17, 21):
            if bool(sheet_t.cell(row=r, column=9).value): reforco_count += 1
            
        tecnica_count = 0
        for r in range(5, 9):
            if bool(sheet_t.cell(row=r, column=27).value): tecnica_count += 1
            
        dom_count = 0
        for r in range(5, 9):
            if bool(sheet_t.cell(row=r, column=45).value): dom_count += 1
        
        resistencia_count = 0
        for r in range(17, 21):
            if bool(sheet_t.cell(row=r, column=45).value): resistencia_count += 1
            
        if not pericias.get('_treinamentos'):
            pericias['_treinamentos'] = {}
            
        pericias['_treinamentos']['reforco_corp'] = min(reforco_count, 4)
        pericias['_treinamentos']['tecnica_ref'] = min(tecnica_count, 4)
        pericias['_treinamentos']['dom_simples'] = min(dom_count, 4)
        pericias['_treinamentos']['resistencia'] = min(resistencia_count, 4)
        
    char.pericias = json.dumps(pericias)
    db.session.commit()
    
    if char.status:
        char.status.pv_atual = char.status.pv_max
        char.status.pe_atual = char.status.pe_max
        char.status.integridade_atual = char.status.integridade_max
        
    db.session.commit()

    import_summary = {
        'atributos': '6 atributos importados',
        'pericias': f'{sum(1 for p in pericias.values() if isinstance(p, dict) and p.get("treinada"))} perícias treinadas detectadas',
        'resistencias': '5 resistências importadas',
        'rds': f'{sum(1 for v in rds.values() if v > 0)} RDs ativas importadas',
        'ataques': f'{len(attacks)} ataque(s) importado(s)' if attacks else 'Nenhum ataque encontrado',
        'talentos': f'{len(talents)} habilidade(s)/talento(s) importado(s)' if talents else 'Nenhum talento encontrado',
        'inventario': f'{len(inventory)} item(ns) de inventário importado(s)' if inventory else 'Nenhum item no inventário',
        'feiticos': f'{len(spells)} feitiço(s) importado(s)' if spells else 'Nenhum feitiço encontrado',
        'summons': f'{len(summons)} invocação(ões) importada(s)' if summons else 'Nenhuma invocação encontrada',
    }

    try:
        import time
        logs = json.loads(char.recent_logs or '[]')
        summary_lines = list(import_summary.values())
        logs.insert(0, {
            'title': 'Ficha Importada!',
            'content': f"Excel '{filename}' importado. " + " | ".join(summary_lines[:3]),
            'time': time.strftime('%H:%M'),
            'timestamp': time.time()
        })
        char.recent_logs = json.dumps(logs[:15])
        db.session.commit()
    except:
        pass

    return import_summary

@app.route('/api/import_excel/<int:character_id>', methods=['POST'])
@login_required
def import_excel(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado.'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nome de arquivo inválido.'}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Falha ao carregar o arquivo Excel: {str(e)}'}), 400

    required_sheets = ['Ficha Pessoal', 'Registro e Inventário', 'Perfil Amaldiçoado']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            return jsonify({'error': f'A planilha não possui a aba obrigatória: "{sheet_name}"'}), 400

    try:
        import_summary = _process_excel_import(char, wb, file.filename)
        return jsonify({
            'message': 'Ficha do Excel importada com total sucesso!',
            'import_summary': import_summary,
            'character': get_character_json(char)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao processar conteúdo da planilha: {str(e)}'}), 400

@app.route('/api/create_character_from_excel', methods=['POST'])
@login_required
def create_character_from_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado.'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nome de arquivo inválido.'}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Falha ao carregar o arquivo Excel: {str(e)}'}), 400

    required_sheets = ['Ficha Pessoal', 'Registro e Inventário', 'Perfil Amaldiçoado']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            return jsonify({'error': f'A planilha não possui a aba obrigatória: "{sheet_name}"'}), 400

    # 1. Cria uma ficha em branco temporária
    char = Character(
        user_id=current_user.id,
        nome="Feiticeiro Planilha",
        origem="Humano",
        especializacao="Feiticeiro de Combate"
    )
    db.session.add(char)
    db.session.commit()

    # 2. Cria os objetos agregados necessários
    status = Status(character_id=char.id)
    attrs = Attributes(character_id=char.id)
    db.session.add(status)
    db.session.add(attrs)
    db.session.commit()

    # 3. Invoca o shared helper
    try:
        import_summary = _process_excel_import(char, wb, file.filename)
        return jsonify({
            'message': 'Personagem criado e importado do Excel com total sucesso!',
            'import_summary': import_summary,
            'character': get_character_json(char)
        })
    except Exception as e:
        db.session.delete(char)
        db.session.commit()
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao processar conteúdo da planilha: {str(e)}'}), 400

@app.route('/api/create_character_from_excel_url', methods=['POST'])
@login_required
def create_character_from_excel_url():
    data = request.json or {}
    url = data.get('url', '').strip()
    if not url:
        return jsonify({'error': 'Por favor, informe a URL da planilha.'}), 400
        
    import re
    match = re.search(r'/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        return jsonify({'error': 'Link da planilha inválido. Certifique-se de que é um link válido do Google Sheets.'}), 400
        
    spreadsheet_id = match.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    
    try:
        import urllib.request
        import urllib.error
        import io
        import openpyxl
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'
        }
        req = urllib.request.Request(export_url, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=30) as response:
                content = response.read()
        except urllib.error.HTTPError as e:
            if e.code in (401, 403):
                return jsonify({'error': 'Acesso negado pela planilha do Google Sheets. Verifique se o compartilhamento está definido como "Qualquer pessoa com o link" (Leitor).'}), 400
            return jsonify({'error': f'Erro HTTP {e.code} ao conectar com o Google Sheets. Verifique o link e o compartilhamento.'}), 400
        except Exception as e:
            return jsonify({'error': f'Falha ao baixar planilha do Google Sheets (problema de rede): {str(e)}'}), 400
            
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return jsonify({'error': f'Falha ao processar a planilha do Google Sheets: {str(e)}'}), 400
        
    required_sheets = ['Ficha Pessoal', 'Registro e Inventário', 'Perfil Amaldiçoado']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            return jsonify({'error': f'A planilha não possui a aba obrigatória: "{sheet_name}"'}), 400
            
    # 1. Cria uma ficha em branco temporária
    char = Character(
        user_id=current_user.id,
        nome="Feiticeiro Planilha",
        origem="Humano",
        especializacao="Feiticeiro de Combate"
    )
    db.session.add(char)
    db.session.commit()

    # 2. Cria os objetos agregados necessários
    status = Status(character_id=char.id)
    attrs = Attributes(character_id=char.id)
    db.session.add(status)
    db.session.add(attrs)
    db.session.commit()

    # 3. Invoca o shared helper
    try:
        import_summary = _process_excel_import(char, wb, "Google Sheets")
        
        # Reset current health & stamina to max
        if char.status:
            char.status.pv_atual = char.status.pv_max
            char.status.pe_atual = char.status.pe_max
            db.session.commit()
            
        return jsonify({
            'message': 'Personagem criado e importado do Google Sheets com total sucesso!',
            'import_summary': import_summary,
            'character': get_character_json(char)
        })
    except Exception as e:
        db.session.delete(char)
        db.session.commit()
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao processar conteúdo da planilha: {str(e)}'}), 400

@app.route('/api/update_attack_jogadas/<int:character_id>', methods=['POST'])
@login_required
def update_attack_jogadas(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.json or {}
    try:
        config = json.loads(char.configuracoes or '{}')
    except:
        config = {}
        
    if 'ataque_corpo_corpo' in data:
        config['ataque_corpo_corpo'] = data['ataque_corpo_corpo']
    if 'ataque_a_distancia' in data:
        config['ataque_a_distancia'] = data['ataque_a_distancia']
    if 'ataque_amaldicoado' in data:
        config['ataque_amaldicoado'] = data['ataque_amaldicoado']
        
    char.configuracoes = json.dumps(config)
    db.session.commit()
    
    return jsonify({
        'message': 'Jogadas de ataque atualizadas com sucesso!',
        'character': get_character_json(char)
    })

@app.route('/api/import_excel_url/<int:character_id>', methods=['POST'])
@login_required
def import_excel_url(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.json or {}
    url = data.get('url', '').strip()
    if not url:
        return jsonify({'error': 'Por favor, informe a URL da planilha.'}), 400
        
    import re
    match = re.search(r'/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        return jsonify({'error': 'Link da planilha inválido. Certifique-se de que é um link válido do Google Sheets.'}), 400
        
    spreadsheet_id = match.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    
    try:
        import urllib.request
        import urllib.error
        import io
        import openpyxl
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'
        }
        req = urllib.request.Request(export_url, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=30) as response:
                content = response.read()
        except urllib.error.HTTPError as e:
            if e.code in (401, 403):
                return jsonify({'error': 'Acesso negado pela planilha do Google Sheets. Verifique se o compartilhamento está definido como "Qualquer pessoa com o link" (Leitor).'}), 400
            return jsonify({'error': f'Erro HTTP {e.code} ao conectar com o Google Sheets. Verifique o link e o compartilhamento.'}), 400
        except Exception as e:
            return jsonify({'error': f'Falha ao baixar planilha do Google Sheets (problema de rede): {str(e)}'}), 400
            
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return jsonify({'error': f'Falha ao processar a planilha do Google Sheets: {str(e)}'}), 400
        
    required_sheets = ['Ficha Pessoal', 'Registro e Inventário', 'Perfil Amaldiçoado']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            return jsonify({'error': f'A planilha não possui a aba obrigatória: "{sheet_name}"'}), 400
            
    try:
        import_summary = _process_excel_import(char, wb, "Google Sheets")
        
        # Reset current health & stamina to max upon successful full import override!
        if char.status:
            char.status.pv_atual = char.status.pv_max
            char.status.pe_atual = char.status.pe_max
            db.session.commit()
            
        return jsonify({
            'message': 'Ficha do Google Sheets importada com total sucesso, substituindo todos os dados anteriores!',
            'import_summary': import_summary,
            'character': get_character_json(char)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao processar conteúdo da planilha: {str(e)}'}), 400

@app.errorhandler(404)
def not_found(e):
    if request.headers.get('Accept') == 'application/json' or request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Not Found'}), 404
    return render_template('index.html')

@app.route('/api/log_action/<int:character_id>', methods=['POST'])
@login_required
def log_action(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    if not data or 'title' not in data or 'content' not in data:
        return jsonify({'error': 'Invalid data'}), 400
        
    try:
        logs = json.loads(char.recent_logs or '[]')
    except:
        logs = []
        
    new_log = {
        'title': data['title'],
        'content': data['content'],
        'time': data.get('time', ''),
        'timestamp': time.time()
    }
    
    logs.insert(0, new_log)
    logs = logs[:15] # keep last 15 items
    
    char.recent_logs = json.dumps(logs)
    db.session.commit()
    
    return jsonify(logs)

@app.route('/api/update_dominio/<int:character_id>', methods=['POST'])
@login_required
def update_dominio(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Invalid data'}), 400
        
    try:
        dominio_data = json.loads(char.dominio or '{}')
    except:
        dominio_data = {}
        
    if 'nome' in data: dominio_data['nome'] = data['nome']
    if 'tipo' in data: dominio_data['tipo'] = data['tipo']
    if 'custo' in data: dominio_data['custo'] = int(data['custo'])
    if 'descricao' in data: dominio_data['descricao'] = data['descricao']
    
    char.dominio = json.dumps(dominio_data)
    db.session.commit()
    
    return jsonify({
        'message': 'Domínio atualizado com sucesso',
        'character': get_character_json(char)
    })

@app.route('/api/manifestar_dominio/<int:character_id>', methods=['POST'])
@login_required
def manifestar_dominio(character_id):
    char = Character.query.get_or_404(character_id)
    if current_user.role == 'Jogador' and char.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    if not char.status:
        return jsonify({'error': 'Character status not initialized'}), 400
        
    try:
        dominio_data = json.loads(char.dominio or '{}')
    except:
        dominio_data = {
            'nome': 'Expansão de Domínio',
            'tipo': 'Letal',
            'custo': 20,
            'descricao': 'A técnica cria um espaço separado fechado por uma barreira e imbuído da técnica inata do usuário. Os ataques dentro de um domínio possuem acerto garantido.'
        }
        
    cost = int(dominio_data.get('custo', 20))
    if char.status.pe_atual < cost:
        return jsonify({'error': 'Pontos de Energia (PE) insuficientes!'}), 400
        
    char.status.pe_atual -= cost
    db.session.commit()
    
    val, desc = roll_dice("1d20+12")
    
    return jsonify({
        'message': f"Manifestou a Expansão de Domínio: {dominio_data['nome']}",
        'dominio': dominio_data,
        'pe_atual': char.status.pe_atual,
        'cost': cost,
        'roll_val': val,
        'roll_desc': desc,
        'character': get_character_json(char)
    })

@app.route('/proxy/owlbear/<path:subpath>', methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])
def proxy_owlbear(subpath):
    import urllib.request
    import urllib.error
    import urllib.parse
    
    # Handle preflight OPTIONS request
    if request.method == 'OPTIONS':
        from flask import Response
        res = Response()
        res.headers['Access-Control-Allow-Origin'] = '*'
        res.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        res.headers['Access-Control-Allow-Headers'] = '*'
        return res
        
    safe_subpath = urllib.parse.quote(subpath, safe='/')
    
    # Smart URL routing for custom domains/subdomains/cloudflare challenges
    if safe_subpath.startswith('data.owlbear.rodeo') or safe_subpath.startswith('challenges.cloudflare.com'):
        target_url = f"https://{safe_subpath}"
    elif '/' in safe_subpath and (safe_subpath.split('/')[0].endswith('.rodeo') or safe_subpath.split('/')[0].endswith('.com') or safe_subpath.split('/')[0].endswith('.app')):
        target_url = f"https://{safe_subpath}"
    else:
        target_url = f"https://www.owlbear.rodeo/{safe_subpath}"
        
    if request.query_string:
        target_url += f"?{request.query_string.decode('utf-8')}"
        
    # Rebuild headers
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'
    }
    
    # Forward common client headers
    for h in ['Accept', 'Accept-Language', 'Content-Type', 'Authorization']:
        if h in request.headers:
            headers[h] = request.headers[h]
            
    try:
        # If there's body data, forward it
        data = None
        if request.method in ['POST', 'PUT', 'DELETE']:
            data = request.get_data()
            
        req = urllib.request.Request(
            target_url, 
            data=data,
            headers=headers,
            method=request.method
        )
        
        with urllib.request.urlopen(req, timeout=30) as res:
            content = res.read()
            status = res.status
            content_type = res.headers.get('Content-Type', 'text/html')
            
            # If it's HTML, inject our monkey-patch proxy script AND rewrite assets
            if 'text/html' in content_type:
                html = content.decode('utf-8', errors='ignore')
                
                # Ultimate CORS & CSP bypass script injection
                proxy_script = r"""
<script>
(function() {
  const origin = window.location.origin;
  const proxyPrefix = origin + '/proxy/owlbear/';

  function toProxyUrl(url) {
    if (!url) return url;
    let urlStr = typeof url === 'string' ? url : url.toString();
    
    if (urlStr.startsWith(proxyPrefix) || urlStr.includes('/proxy/owlbear/')) {
      return url;
    }
    if (urlStr.startsWith('/') && !urlStr.startsWith('//')) {
      return urlStr;
    }
    if (urlStr.includes('owlbear.rodeo') || urlStr.includes('owlbear.app') || urlStr.includes('cloudflare.com')) {
      const cleanUrl = urlStr.replace(/https?:\/\//, '');
      return '/proxy/owlbear/' + cleanUrl;
    }
    return url;
  }

  const originalFetch = window.fetch;
  window.fetch = function(input, init) {
    if (typeof input === 'string') {
      input = toProxyUrl(input);
    } else if (input instanceof Request) {
      const proxyUrl = toProxyUrl(input.url);
      input = new Request(proxyUrl, input);
    }
    return originalFetch.apply(this, [input, init]);
  };

  const originalOpen = XMLHttpRequest.prototype.open;
  XMLHttpRequest.prototype.open = function(method, url, async, user, password) {
    const proxyUrl = toProxyUrl(url);
    return originalOpen.apply(this, [method, proxyUrl, async, user, password]);
  };
})();
</script>
"""
                
                if '<head>' in html:
                    html = html.replace('<head>', f'<head>{proxy_script}', 1)
                elif '<HEAD>' in html:
                    html = html.replace('<HEAD>', f'<HEAD>{proxy_script}', 1)
                else:
                    html = f"{proxy_script}{html}"
                
                # Rewrite all absolute paths to go through our proxy
                html = html.replace('href="/', 'href="/proxy/owlbear/')
                html = html.replace('src="/', 'src="/proxy/owlbear/')
                html = html.replace("href='/", "href='/proxy/owlbear/")
                html = html.replace("src='/", "src='/proxy/owlbear/")
                html = html.replace('"/manifest.json"', '"/proxy/owlbear/manifest.json"')
                html = html.replace("'/manifest.json'", "'/proxy/owlbear/manifest.json'")
                    
                content = html.encode('utf-8')
                
            from flask import Response
            response = Response(content, status=status)
            response.headers['Content-Type'] = content_type
            
            # Explicitly allow framing and CORS for all requests
            response.headers['X-Frame-Options'] = 'ALLOWALL'
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
            response.headers['Access-Control-Allow-Headers'] = '*'
            
            return response
            
    except urllib.error.HTTPError as e:
        try:
            content = e.read()
        except:
            content = b"Error loading resource"
        from flask import Response
        res = Response(content, status=e.code)
        if 'Content-Type' in e.headers:
            res.headers['Content-Type'] = e.headers['Content-Type']
        res.headers['X-Frame-Options'] = 'ALLOWALL'
        res.headers['Access-Control-Allow-Origin'] = '*'
        return res
    except Exception as e:
        return jsonify({'error': f'Proxy Error: {str(e)}'}), 500

@app.route('/assets/<path:path>', methods=['GET'])
def proxy_assets(path):
    import urllib.request
    import urllib.error
    
    target_url = f"https://www.owlbear.rodeo/assets/{path}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'
    }
    
    try:
        req = urllib.request.Request(target_url, headers=headers)
        with urllib.request.urlopen(req, timeout=30) as res:
            content = res.read()
            status = res.status
            content_type = res.headers.get('Content-Type', '')
            
            from flask import Response
            response = Response(content, status=status)
            if content_type:
                response.headers['Content-Type'] = content_type
                
            response.headers['X-Frame-Options'] = 'ALLOWALL'
            response.headers['Access-Control-Allow-Origin'] = '*'
            return response
    except Exception as e:
        return f"Asset proxy error: {str(e)}", 404

@app.route('/room/<path:subpath>', methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])
def proxy_room(subpath):
    return proxy_owlbear(f"room/{subpath}")

@app.route('/sign-up', methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])
def proxy_signup():
    return proxy_owlbear("sign-up")

@app.route('/manifest.json', methods=['GET'])
def proxy_manifest():
    return proxy_owlbear("manifest.json")

@app.route('/cdn-cgi/<path:path>', methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])
def proxy_cdn_cgi(path):
    return proxy_owlbear(f"cdn-cgi/{path}")

if __name__ == '__main__':
    # Ensure templates and static folders exist
    os.makedirs(os.path.join(base_dir, 'templates'), exist_ok=True)
    os.makedirs(os.path.join(base_dir, 'static'), exist_ok=True)
    
    port = int(os.environ.get('FLASK_PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
