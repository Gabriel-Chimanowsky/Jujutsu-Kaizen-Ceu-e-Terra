import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet = wb[wb.sheetnames[4]]

print("--- SUMMONS TOP VALUES (Rows 5-15) ---")
for r in range(5, 16):
    row_vals = []
    for c in range(2, 16): # B(2) to O(15)
        val = sheet.cell(row=r, column=c).value
        coord = sheet.cell(row=r, column=c).coordinate
        row_vals.append(f"{coord}:{val}")
    print(f"Row {r}: {row_vals}")
