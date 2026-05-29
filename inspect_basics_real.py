import openpyxl

wb_v = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet_v = wb_v['Ficha Pessoal']

wb_f = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet_f = wb_f['Ficha Pessoal']

coords = ['S2', 'AH2', 'AH3', 'AU3', 'AU4', 'BD3', 'AA7', 'AA8', 'AA9', 'AA10', 'Z11']

print("--- RAW & EVALUATED VALUES ---")
for coord in coords:
    val_v = sheet_v[coord].value
    val_f = sheet_f[coord].value
    print(f"{coord} -> Value: {val_v} | Formula: {val_f}")

print("\n--- ALL NON-EMPTY CELLS IN FIRST 5 ROWS ---")
for r in range(1, 6):
    for c in range(1, 80):
        val_v = sheet_v.cell(row=r, column=c).value
        if val_v is not None:
            coord = sheet_v.cell(row=r, column=c).coordinate
            print(f"{coord}: {val_v}")
