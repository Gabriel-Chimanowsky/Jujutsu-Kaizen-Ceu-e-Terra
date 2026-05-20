import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=False)
sheet = wb['Ficha Pessoal']

print("--- Row 27 non-empty cells ---")
for c in range(1, 100):
    val = sheet.cell(row=27, column=c).value
    if val is not None:
        print(f"Col {c} ({sheet.cell(row=27, column=c).coordinate}): {val}")
