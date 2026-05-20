import openpyxl

wb = openpyxl.load_workbook("exemplo/Modelo de Ficha - Feiticeiros & Maldições 2.5.xlsx", data_only=True)
sheet = wb['Ficha Pessoal']

print("--- ATTACK DETAILS (Rows 27-31) ---")
for r in range(27, 32):
    arma = sheet.cell(row=r, column=27).value # AA is 27
    bonus = sheet.cell(row=r, column=32).value # AF is 32
    dano = sheet.cell(row=r, column=34).value # AH is 34
    crit = sheet.cell(row=r, column=38).value # AL is 38
    tipo = sheet.cell(row=r, column=40).value # AN is 40
    alcance = sheet.cell(row=r, column=43).value # AQ is 43
    print(f"Row {r} -> ARMA: {arma}, BÔNUS: {bonus}, DANO: {dano}, CRÍT: {crit}, TIPO: {tipo}, ALCANCE: {alcance}")
